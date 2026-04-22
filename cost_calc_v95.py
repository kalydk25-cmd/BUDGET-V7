# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import platform
import re
import math
import datetime
import sys
import csv
from collections import defaultdict

try:
    from fixed_staff_wage_data import FIXED_WAGE_SOURCE, FIXED_WAGE_RECORDS, FIXED_PERSON_MASTER
except Exception:
    FIXED_WAGE_SOURCE = "人力版本日薪及人岗结构汇总.xlsx"
    FIXED_WAGE_RECORDS = []
    FIXED_PERSON_MASTER = []

# 尝试导入 pandas
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# 尝试导入 openpyxl (用于精美 xlsx 报表)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ==========================================
# 1. 工具函数与核心逻辑
# ==========================================
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

def load_json(file, default):
    if os.path.exists(file):
        try: return json.load(open(file, "r", encoding="utf-8"))
        except: return default.copy()
    else: return default.copy()

def save_json(file, data):
    try: json.dump(data, open(file, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    except: pass

def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", os.path.abspath(os.path.dirname(__file__)))
    return os.path.join(base_path, relative_path)

def write_csv(file_path, headers, rows):
    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


EMERGENCY_RESERVE_RATE = 0.10

def export_styled_xlsx(file_path, summary_data, detail_data, title, config_data=None):
    wb = Workbook()

    # ---- 封面 ----
    ws_cover = wb.active
    ws_cover.title = "封面"
    ws_cover.sheet_view.showGridLines = False
    ws_cover.column_dimensions["A"].width = 6
    ws_cover.column_dimensions["B"].width = 14
    ws_cover.column_dimensions["C"].width = 14
    ws_cover.column_dimensions["D"].width = 14
    ws_cover.column_dimensions["E"].width = 14
    ws_cover.column_dimensions["F"].width = 14
    ws_cover.column_dimensions["G"].width = 14
    ws_cover.column_dimensions["H"].width = 14
    ws_cover.row_dimensions[2].height = 60
    ws_cover.row_dimensions[6].height = 32
    ws_cover.row_dimensions[8].height = 22
    ws_cover.row_dimensions[10].height = 26

    ws_cover.merge_cells("B6:H6")
    ws_cover["B6"] = title
    ws_cover["B6"].font = Font(name="Microsoft YaHei UI", size=20, bold=True, color="1F4CC4")
    ws_cover["B6"].alignment = Alignment(horizontal="center", vertical="center")

    ws_cover.merge_cells("B8:H8")
    ws_cover["B8"] = f"生成日期：{datetime.datetime.now().strftime('%Y-%m-%d')}"
    ws_cover["B8"].font = Font(name="Microsoft YaHei UI", size=11, color="667085")
    ws_cover["B8"].alignment = Alignment(horizontal="center", vertical="center")

    total_value = None
    for row in summary_data:
        if row.get("项目") == "总成本预估":
            total_value = row.get("金额")
            break
    if total_value is not None:
        ws_cover.merge_cells("B10:H10")
        ws_cover["B10"] = f"总成本预估：¥{total_value:,.0f}"
        ws_cover["B10"].font = Font(name="Microsoft YaHei UI", size=16, bold=True, color="E11D48")
        ws_cover["B10"].alignment = Alignment(horizontal="center", vertical="center")

    logo_path = resource_path("LOGO.png")
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            max_w = 180
            if img.width and img.width > max_w:
                scale = max_w / img.width
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
            ws_cover.add_image(img, "D2")
        except:
            pass

    # ---- 总价汇总表 ----
    ws_sum = wb.create_sheet("总价汇总表")
    headers_sum = ["项目", "金额", "备注"]

    header_fill = PatternFill("solid", fgColor="2F6BFF")
    header_font = Font(name="Microsoft YaHei UI", color="FFFFFF", bold=True)
    zebra_a = PatternFill("solid", fgColor="FFFFFF")
    zebra_b = PatternFill("solid", fgColor="F5F7FB")
    total_fill = PatternFill("solid", fgColor="FFF4CC")
    thin = Side(style="thin", color="D6DFEA")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    ws_sum.append(headers_sum)
    for col in range(1, len(headers_sum) + 1):
        cell = ws_sum.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border

    for idx, row in enumerate(summary_data, start=2):
        ws_sum.append([row.get("项目"), row.get("金额"), row.get("备注")])
        is_total = row.get("项目") == "总成本预估"
        fill = total_fill if is_total else (zebra_b if idx % 2 == 0 else zebra_a)
        for col in range(1, 4):
            cell = ws_sum.cell(row=idx, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = align_center if col != 2 else align_right
            if col == 2:
                cell.number_format = "¥#,##0"
                if is_total:
                    cell.font = Font(name="Microsoft YaHei UI", bold=True)

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 18
    ws_sum.column_dimensions["C"].width = 40
    ws_sum.freeze_panes = "A2"

    # ---- 人工明细表 ----
    ws_det = wb.create_sheet("人工明细表")
    headers_det = ["任务", "专业", "工日", "工时"]
    ws_det.append(headers_det)
    for col in range(1, len(headers_det) + 1):
        cell = ws_det.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border

    for idx, row in enumerate(detail_data, start=2):
        ws_det.append([row.get("task"), row.get("dept"), row.get("work_days"), row.get("work_hours")])
        fill = zebra_b if idx % 2 == 0 else zebra_a
        for col in range(1, 5):
            cell = ws_det.cell(row=idx, column=col)
            cell.fill = fill
            cell.border = border
            if col in (3, 4):
                cell.number_format = "0.0"
                cell.alignment = align_center
            else:
                cell.alignment = align_center

    ws_det.column_dimensions["A"].width = 36
    ws_det.column_dimensions["B"].width = 15
    ws_det.column_dimensions["C"].width = 10
    ws_det.column_dimensions["D"].width = 10
    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = f"A1:D{ws_det.max_row}"

    # ---- 参数与选项 ----
    ws_cfg = wb.create_sheet("参数与选项")
    headers_cfg = ["分类", "项目", "值"]
    ws_cfg.append(headers_cfg)
    for col in range(1, len(headers_cfg) + 1):
        cell = ws_cfg.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border

    for idx, row in enumerate(config_data or [], start=2):
        ws_cfg.append([row.get("分类"), row.get("项目"), row.get("值")])
        fill = zebra_b if idx % 2 == 0 else zebra_a
        for col in range(1, 4):
            cell = ws_cfg.cell(row=idx, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = align_center if col != 3 else Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_cfg.column_dimensions["A"].width = 18
    ws_cfg.column_dimensions["B"].width = 28
    ws_cfg.column_dimensions["C"].width = 90
    ws_cfg.freeze_panes = "A2"
    ws_cfg.auto_filter.ref = f"A1:C{ws_cfg.max_row}"

    wb.save(file_path)

def parse_rules(rule_str):
    if not rule_str: return []
    rules = []
    try:
        clean_str = str(rule_str).replace('MW','').replace('mw','').replace('×','*').replace('：',':').replace('~','-')
        for line in clean_str.split('\n'):
            if '*' not in line: continue
            parts = line.split('*')
            cond = parts[0].strip()
            try: factor = float(parts[1])
            except: continue
            nums = [float(x) for x in re.findall(r"[\d\.]+", cond)]
            if not nums: continue
            min_v, max_v = float('-inf'), float('inf')
            if '≤' in cond or '<=' in cond: max_v = nums[0]
            elif '≥' in cond or '>=' in cond: min_v = nums[0]
            elif '+' in cond: min_v = nums[0]
            elif '-' in cond and len(nums) >= 2: min_v, max_v = nums[0], nums[1]
            elif len(nums) == 1: max_v = nums[0]
            rules.append((min_v, max_v, factor))
    except: pass
    rules.sort(key=lambda x: x[1])
    return rules

def get_capacity_factor(capacity, rules):
    if not rules: return 1.0
    for min_v, max_v, factor in rules:
        if capacity >= min_v and capacity <= max_v: return factor
    for min_v, max_v, factor in rules:
        if capacity <= max_v: return factor
    if rules: return rules[-1][2]
    return 1.0

def normalize_matrix(matrix_data):
    corrected_matrix = {}
    for task_name, depts in matrix_data.items():
        total_ratio = sum(depts.values())
        if total_ratio <= 0: 
            corrected_matrix[task_name] = depts
            continue
        if abs(total_ratio - 100.0) < 0.0001:
            corrected_matrix[task_name] = depts
        else:
            scale_factor = 100.0 / total_ratio
            new_depts = {k: v * scale_factor for k, v in depts.items()}
            corrected_matrix[task_name] = new_depts
    return corrected_matrix

# ==========================================
# 2. 全局数据字典
# ==========================================

WAGE_DATABASE = {
    "项目设总": {
        "主设人": 1400,
        "设计人": 1400,
        "校核人": 1400,
        "审核人": 1400
    },
    "资源": {
        "主设人": 840,
        "设计人": 840,
        "校核人": 1050,
        "审核人": 1330
    },
    "技经": {
        "主设人": 770,
        "设计人": 770,
        "校核人": 1190,
        "审核人": 1400
    },
    "综合能源": {
        "主设人": 770,
        "设计人": 770,
        "校核人": 1190,
        "审核人": 1400
    },
    "电气一次": {
        "主设人": 840,
        "设计人": 840,
        "校核人": 1190,
        "审核人": 1680
    },
    "电气二次": {
        "主设人": 910,
        "设计人": 910,
        "校核人": 1190,
        "审核人": 1400
    },
    "线路电气": {
        "主设人": 980,
        "设计人": 980,
        "校核人": 1330,
        "审核人": 1540
    },
    "线路结构": {
        "主设人": 910,
        "设计人": 910,
        "校核人": 1400,
        "审核人": 1400
    },
    "土建结构": {
        "主设人": 910,
        "设计人": 910,
        "校核人": 1190,
        "审核人": 1540
    },
    "土建建筑": {
        "主设人": 910,
        "设计人": 910,
        "校核人": 910,
        "审核人": 910
    },
    "总图": {
        "主设人": 770,
        "设计人": 770,
        "校核人": 1050,
        "审核人": 1680
    },
    "道路": {
        "主设人": 910,
        "设计人": 910,
        "校核人": 1120,
        "审核人": 1400
    },
    "水暖": {
        "主设人": 910,
        "设计人": 910,
        "校核人": 910,
        "审核人": 1190
    },
    "岩土": {
        "主设人": 910,
        "设计人": 910,
        "校核人": 910,
        "审核人": 1400
    },
    "测绘": {
        "主设人": 700,
        "设计人": 700,
        "校核人": 980,
        "审核人": 1190
    },
    "DEFAULT": {
        "主设人": 840,
        "设计人": 840,
        "校核人": 1050,
        "审核人": 1330
    }
}

DEFAULT_RATIOS = {"DEFAULT": {"主设人": 5, "设计人": 71, "校核人": 16, "审核人": 8}}

PROCUREMENT_DEFAULTS = {
    "风电": {"升压站基准": {"初勘": 40, "详勘": 250}},
    "光伏": {"孔深": 7, "升压站基准": {"初勘": 40, "详勘": 250}, "单价": {"初勘": 130, "详勘": 110}},
    "储能": {"孔深": 7, "升压站基准": {"初勘": 40, "详勘": 250}, "单价": {"初勘": 130, "详勘": 110}}
}

# 设计院其他费用成本标准（2026年版），单位：万元
OTHER_COST_STD = {
    "consult_items_wan": {
        "traffic": {"le_100": 0.88, "101_200": 0.88, "201_500": 0.88, "gt_500": 0.88},
        "lodging": {"le_100": 0.288, "101_200": 0.378, "201_500": 0.468, "gt_500": 0.648},
        "rental": {"le_100": 0.325, "101_200": 0.4875, "201_500": 0.65, "gt_500": 0.975},
        "subsidy": {"le_100": 0.16, "101_200": 0.24, "201_500": 0.32, "gt_500": 0.48},
        "printing": {"le_100": 0.30, "101_200": 0.30, "201_500": 0.30, "gt_500": 0.30},
    },
    "design_items_wan": {
        "traffic": 2.09,
        "lodging_le_200": 0.936,
        "lodging_gt_200": 1.88,
        "subsidy_le_200": 0.52,
        "subsidy_gt_200": 1.00,
        "printing_domestic": 1.80,
        "printing_international": 2.60,
        "stationed_le_200": 4.00,
        "stationed_gt_200": 5.00,
    },
}

# 废弃旧的阶梯，使用新逻辑，这里仅保留变量定义防止报错，实际逻辑在类里重写
SURVEY_TIERS = [] 

DISCIPLINE_ORDER = [
    "项目设总", "资源", "技经", "综合能源", "电气一次", "电气二次", "线路电气", "线路结构", 
    "土建结构", "土建建筑", "总图", "道路", "水暖", "岩土", "测绘"
]

TASK_SYNONYMS = {
    "可研": ["可行性研究（风电项目，含升压站）", "可行性研究（升压电站）", "可行性研究（光伏项目，含升压站）", "可行性研究（储能电站）", "可研（申报版）、预可研"],
    "初设": ["初步设计报告（风电项目，含升压站）", "初步设计报告（升压电站）", "初步设计报告（光伏项目，含升压站）", "初步设计报告（储能电站）"],
    "施工图": ["施工图（风电项目，含升压站）", "施工图（升压电站）", "施工图（光伏项目，含升压站）", "施工图（储能电站）"],
    "竣工图": ["竣工图（风电项目，含升压站）", "竣工图（光伏项目，含升压站）"],
    "微观": "微观选址报告（风电）",
    "建议书": ["项目建议书（申报版）", "项目建议书（正式版）、项目申请报告"],
    "投标": ["投标技术文件 (咨询)", "投标技术文件 (勘察)", "投标技术文件 (EPC)"],
    "测风": "测风测光方案"
}

RAW_MATRIX_WIND = {
    '测风测光方案': {'资源': 100.0},
    '测风测光年报': {'资源': 100.0},
    '投标技术文件 (咨询)': {'项目设总': 48.78, '资源': 14.63, '技经': 4.88, '电气一次': 7.32, '电气二次': 4.88, '线路电气': 3.9, '线路结构': 0.98, '总图': 2.44, '道路': 2.44, '土建结构': 4.88, '土建建筑': 4.88},
    '投标技术文件 (勘察)': {'项目设总': 38.83, '资源': 11.65, '技经': 11.65, '电气一次': 5.83, '电气二次': 5.83, '线路电气': 3.88, '总图': 3.88, '道路': 3.88, '土建结构': 5.83, '土建建筑': 5.83, '水暖': 2.91},
    '投标技术文件 (EPC)': {'项目设总': 27.03, '资源': 9.46, '技经': 14.86, '电气一次': 9.46, '电气二次': 5.41, '线路电气': 9.46, '总图': 4.05, '道路': 4.05, '土建结构': 6.76, '土建建筑': 6.76, '水暖': 2.7},
    '项目建议书（申报版）': {'项目设总': 25.0, '资源': 50.0, '技经': 25.0},
    '项目建议书（正式版）、项目申请报告': {'项目设总': 25.0, '资源': 25.0, '技经': 8.33, '电气一次': 12.5, '线路电气': 8.33, '总图': 4.17, '道路': 8.33, '土建结构': 4.17, '水暖': 4.17},
    '可研（申报版）、预可研': {'项目设总': 19.05, '资源': 23.81, '技经': 9.52, '电气一次': 19.05, '总图': 4.76, '道路': 9.52, '土建结构': 9.52, '水暖': 4.76},
    '初步规划报告': {'项目设总': 20.0, '资源': 26.67, '技经': 6.67, '电气一次': 20.0, '线路电气': 3.33, '总图': 10.0, '道路': 13.33},
    '第三方报告（资源评估报告）': {'项目设总': 20.0, '资源': 80.0},
    '机型比选报告（风电）': {'项目设总': 12.9, '资源': 38.71, '技经': 25.81, '电气一次': 6.45, '线路电气': 6.45, '道路': 6.45, '土建结构': 3.23},
    '微观选址报告（风电）': {'项目设总': 9.09, '资源': 54.55, '线路电气': 9.09, '总图': 9.09, '道路': 18.18},
    '技术尽调（未建项目）': {'项目设总': 34.78, '资源': 34.78, '技经': 8.7, '电气一次': 13.04, '道路': 4.35, '土建结构': 4.35},
    '项目审查': {'项目设总': 10.0, '资源': 10.0, '技经': 10.0, '电气一次': 10.0, '线路电气': 10.0, '总图': 10.0, '道路': 10.0, '土建结构': 10.0, '水暖': 10.0, '岩土': 10.0},
    '可行性研究（风电）': {'项目设总': 13.33, '资源': 20.74, '技经': 11.85, '电气一次': 7.41, '电气二次': 5.93, '线路电气': 5.93, '线路结构': 2.22, '总图': 5.93, '道路': 10.37, '土建结构': 8.89, '土建建筑': 2.96, '水暖': 2.22, '岩土': 2.22},
    '后评估报告、技术尽调报告（已建项目）': {'资源': 42.86, '电气一次': 14.29, '电气二次': 14.29, '线路电气': 10.71, '土建结构': 17.86},
    '投资机会研究报告': {'资源': 60.0, '技经': 20.0, '电气一次': 20.0},
    '初步设计报告（风电）': {'项目设总': 18.75, '资源': 12.5, '技经': 10.0, '电气一次': 10.0, '电气二次': 7.5, '线路电气': 6.25, '线路结构': 2.5, '总图': 6.25, '道路': 10.0, '土建结构': 7.5, '土建建筑': 3.75, '水暖': 2.5, '岩土': 2.5},
    '施工图（风电）': {'项目设总': 8.46, '电气一次': 16.93, '电气二次': 13.54, '线路电气': 10.16, '线路结构': 5.08, '总图': 6.77, '道路': 10.97, '土建结构': 16.93, '土建建筑': 6.09, '水暖': 5.08},
    '竣工图（风电）': {'项目设总': 11.74, '电气一次': 9.39, '电气二次': 9.39, '线路电气': 5.63, '线路结构': 3.76, '总图': 7.51, '道路': 18.78, '土建结构': 18.78, '土建建筑': 9.39, '水暖': 5.63},
    '初勘': {'岩土': 100.0}, '详勘': {'岩土': 100.0},
    '可研-地形图测绘(外委)': {'测绘': 100.0}, '可研-人工测量(小区域)': {'测绘': 100.0},
    '初设/施工图-地形图测绘': {'测绘': 100.0}, '初设/施工图-线路测量': {'测绘': 100.0}
}

DATA_WIND_PARAMS = [
    {'key': '测风测光方案', 'base_days': 0.5, 'rule': None, 'intl': True},
    {'key': '测风测光年报', 'base_days': 0.5, 'rule': None, 'intl': True},
    {'key': '投标技术文件 (咨询)', 'base_days': 2.25, 'rule': None, 'intl': True},
    {'key': '投标技术文件 (勘察)', 'base_days': 5.35, 'rule': None, 'intl': True},
    {'key': '投标技术文件 (EPC)', 'base_days': 7.8, 'rule': None, 'intl': True},
    {'key': '项目建议书（申报版）', 'base_days': 4.0, 'rule': None, 'intl': False},
    {'key': '项目建议书（正式版）、项目申请报告', 'base_days': 13.0, 'rule': None, 'intl': False},
    {'key': '可研（申报版）、预可研', 'base_days': 11.0, 'rule': None, 'intl': False},
    {'key': '初步规划报告', 'base_days': 15.0, 'rule': None, 'intl': True},
    {'key': '第三方报告（资源评估报告）', 'base_days': 5.0, 'rule': None, 'intl': True},
    {'key': '机型比选报告（风电）', 'base_days': 15.5, 'rule': None, 'intl': True},
    {'key': '微观选址报告（风电）', 'base_days': 11.0, 'rule': '100MW×1\n101-300MW×1.2\n301-500MW×1.4\n501MW+×1.6', 'intl': True},
    {'key': '技术尽调（未建项目）', 'base_days': 11.5, 'rule': None, 'intl': False},
    {'key': '项目审查', 'base_days': 5.0, 'rule': None, 'intl': False},
    {'key': '可行性研究（风电）', 'base_days': 73.5, 'rule': '100MW×1\n101~200MW×1.2\n201~499MW×1.4\n500~1000MW×1.8\n1000+MW×2.5', 'intl': True},
    {'key': '后评估报告、技术尽调报告（已建项目）', 'base_days': 28.0, 'rule': None, 'intl': True},
    {'key': '投资机会研究报告', 'base_days': 10.0, 'rule': None, 'intl': True},
    {'key': '初步设计报告（风电）', 'base_days': 82.0, 'rule': '≤100MW×1\n101-200MW×1.2\n201-499MW×1.4\n500-1000MW×1.8\n1000+MW×2.5', 'intl': True},
    {'key': '施工图（风电）', 'base_days': 331.0, 'rule': '≤50MW×0.8\n50-100MW×1\n101-200MW×1.2\n201-499MW×1.4\n500-1000MW×1.8\n1001+MW×2.5', 'intl': True},
    {'key': '竣工图（风电）', 'base_days': 9.0, 'rule': '≤50MW×0.5\n50+MW×1', 'intl': True},
    {"key": "初勘", "base_days": 13.0, "rule": None, "intl": True},
    {"key": "详勘", "base_days": 19.0, "rule": None, "intl": True},
    {"key": "可研-地形图测绘(外委)", "base_days": 7.0, "rule": None, "intl": True},
    {"key": "可研-人工测量(小区域)", "base_days": 5.0, "rule": None, "intl": True},
    {"key": "初设/施工图-地形图测绘", "base_days": 7.0, "rule": None, "intl": True},
    {"key": "初设/施工图-线路测量", "base_days": 20.0, "rule": None, "intl": True}
]

RAW_MATRIX_PV = {
    '测风测光方案': {'资源': 100.0},
    '测风测光年报': {'资源': 100.0},
    '投标技术文件 (咨询)': {'项目设总': 48.78, '资源': 14.63, '技经': 4.88, '电气一次': 7.32, '电气二次': 4.88, '线路电气': 3.9, '线路结构': 0.98, '总图': 2.44, '道路': 2.44, '土建结构': 4.88, '土建建筑': 4.88},
    '投标技术文件 (勘察)': {'项目设总': 38.83, '资源': 11.65, '技经': 11.65, '电气一次': 5.83, '电气二次': 5.83, '线路电气': 3.88, '总图': 3.88, '道路': 3.88, '土建结构': 5.83, '土建建筑': 5.83, '水暖': 2.91},
    '投标技术文件 (EPC)': {'项目设总': 27.03, '资源': 9.46, '技经': 14.86, '电气一次': 9.46, '电气二次': 5.41, '线路电气': 9.46, '总图': 4.05, '道路': 4.05, '土建结构': 6.76, '土建建筑': 6.76, '水暖': 2.7},
    '项目建议书（申报版）': {'项目设总': 25.0, '资源': 50.0, '技经': 25.0},
    '项目建议书（正式版）、项目申请报告': {'项目设总': 25.0, '资源': 25.0, '技经': 8.33, '电气一次': 12.5, '线路电气': 8.33, '总图': 4.17, '道路': 8.33, '土建结构': 4.17, '水暖': 4.17},
    '可研（申报版）、预可研': {'项目设总': 19.05, '资源': 23.81, '技经': 9.52, '电气一次': 19.05, '总图': 4.76, '道路': 9.52, '土建结构': 9.52, '水暖': 4.76},
    '初步规划报告': {'项目设总': 20.0, '资源': 26.67, '技经': 6.67, '电气一次': 20.0, '线路电气': 3.33, '总图': 10.0, '道路': 13.33},
    '第三方报告（资源评估报告）': {'项目设总': 20.0, '资源': 80.0},
    '技术尽调（未建项目）': {'项目设总': 34.78, '资源': 34.78, '技经': 8.7, '电气一次': 13.04, '道路': 4.35, '土建结构': 4.35},
    '项目审查': {'项目设总': 10.0, '资源': 10.0, '技经': 10.0, '电气一次': 10.0, '线路电气': 10.0, '总图': 10.0, '道路': 10.0, '土建结构': 10.0, '水暖': 10.0, '岩土': 10.0},
    '可行性研究（光伏）': {'项目设总': 13.95, '资源': 18.6, '技经': 12.4, '电气一次': 7.75, '电气二次': 6.2, '线路电气': 6.2, '线路结构': 2.33, '总图': 6.2, '道路': 9.3, '土建结构': 9.3, '土建建筑': 3.1, '水暖': 2.33, '岩土': 2.33},
    '后评估报告、技术尽调报告（已建项目）': {'资源': 42.86, '电气一次': 14.29, '电气二次': 14.29, '线路电气': 10.71, '土建结构': 17.86},
    '投资机会研究报告': {'资源': 60.0, '技经': 20.0, '电气一次': 20.0},
    '初步设计报告（光伏）': {'项目设总': 18.59, '资源': 12.39, '技经': 9.37, '电气一次': 9.37, '电气二次': 7.49, '线路电气': 6.2, '线路结构': 1.24, '总图': 6.2, '道路': 9.91, '土建结构': 9.91, '土建建筑': 3.72, '水暖': 2.81, '岩土': 2.81},
    '施工图（光伏）': {'项目设总': 8.45, '资源': 10.14, '电气一次': 16.22, '电气二次': 13.51, '线路电气': 6.76, '线路结构': 5.07, '总图': 6.76, '道路': 6.76, '土建结构': 15.2, '土建建筑': 6.08, '水暖': 5.07},
    '竣工图（光伏）': {'项目设总': 10.48, '资源': 14.68, '电气一次': 11.74, '电气二次': 11.74, '线路电气': 4.4, '线路结构': 2.94, '总图': 5.87, '道路': 11.74, '土建结构': 14.68, '土建建筑': 7.34, '水暖': 4.4},
    '初勘': {'岩土': 100.0}, '详勘': {'岩土': 100.0},
    '可研-地形图测绘(外委)': {'测绘': 100.0}, '可研-人工测量(小区域)': {'测绘': 100.0},
    '初设/施工图-地形图测绘': {'测绘': 100.0}, '初设/施工图-线路测量': {'测绘': 100.0}
}

DATA_PV_PARAMS = [
    {'key': '测风测光方案', 'base_days': 0.5, 'rule': None, 'intl': True},
    {'key': '测风测光年报', 'base_days': 0.5, 'rule': None, 'intl': True},
    {'key': '投标技术文件 (咨询)', 'base_days': 2.25, 'rule': None, 'intl': True},
    {'key': '投标技术文件 (勘察)', 'base_days': 5.35, 'rule': None, 'intl': True},
    {'key': '投标技术文件 (EPC)', 'base_days': 7.8, 'rule': None, 'intl': True},
    {'key': '项目建议书（申报版）', 'base_days': 4.0, 'rule': None, 'intl': False},
    {'key': '项目建议书（正式版）、项目申请报告', 'base_days': 13.0, 'rule': None, 'intl': False},
    {'key': '可研（申报版）、预可研', 'base_days': 11.0, 'rule': None, 'intl': False},
    {'key': '初步规划报告', 'base_days': 15.0, 'rule': None, 'intl': True},
    {'key': '第三方报告（资源评估报告）', 'base_days': 5.0, 'rule': None, 'intl': True},
    {'key': '技术尽调（未建项目）', 'base_days': 11.5, 'rule': None, 'intl': False},
    {'key': '项目审查', 'base_days': 5.0, 'rule': None, 'intl': False},
    {'key': '可行性研究（光伏）', 'base_days': 71.0, 'rule': '≤100MW×1\n101-200MW×1.2\n201-499MW×1.4\n500-1000MW×1.8\n1001+MW×2.5', 'intl': True},
    {'key': '后评估报告、技术尽调报告（已建项目）', 'base_days': 28.0, 'rule': None, 'intl': True},
    {'key': '投资机会研究报告', 'base_days': 10.0, 'rule': None, 'intl': True},
    {'key': '初步设计报告（光伏）', 'base_days': 81.704, 'rule': '≤100MW×1\n101-200MW×1.2\n201-499MW×1.4\n500-1000MW×1.8\n1000+MW×2.5', 'intl': True},
    {'key': '施工图（光伏）', 'base_days': 339.0, 'rule': '≤50MW×0.8\n50-100MW×1\n101-200MW×1.2\n201-499MW×1.4\n500-1000MW×1.8\n1001+MW×2.5', 'intl': True},
    {'key': '竣工图（光伏）', 'base_days': 40.44, 'rule': '≤50MW×0.5\n50+MW×1', 'intl': True},
    {"key": "初勘", "base_days": 13.0, "rule": None, "intl": True},
    {"key": "详勘", "base_days": 19.0, "rule": None, "intl": True},
    {"key": "可研-地形图测绘(外委)", "base_days": 7.0, "rule": None, "intl": True},
    {"key": "可研-人工测量(小区域)", "base_days": 5.0, "rule": None, "intl": True},
    {"key": "初设/施工图-地形图测绘", "base_days": 7.0, "rule": None, "intl": True},
    {"key": "初设/施工图-线路测量", "base_days": 20.0, "rule": None, "intl": True}
]

MATRIX_WIND = normalize_matrix(RAW_MATRIX_WIND)
MATRIX_PV = normalize_matrix(RAW_MATRIX_PV)

RATIO_FILE = "dept_ratios.json"
TASK_CONFIG_FILE = "task_config_overrides.json"
CURRENT_STAFF_DB = {} 
CURRENT_RATIO_DB = load_json(RATIO_FILE, DEFAULT_RATIOS)

DB_WIND_PARAMS = {x['key']: {**x, 'parsed_rule': parse_rules(x['rule'])} for x in DATA_WIND_PARAMS}
DB_PV_PARAMS = {x['key']: {**x, 'parsed_rule': parse_rules(x['rule'])} for x in DATA_PV_PARAMS}
CURRENT_TASK_DETAILS = {}

RATIO_COLUMN_MAP = {
    "项目设总": "项目设总",
    "资源": "资源",
    "技经": "技经",
    "电气一次": "电气一次",
    "电气二次": "电气二次",
    "线路电气": "线路电气",
    "线路结构": "线路结构",
    "总图": "总图",
    "道路": "道路",
    "土建结构": "土建结构",
    "土建建筑": "土建建筑",
    "水暖": "水暖",
    "岩土": "岩土",
    "测绘": "测绘",
}

SYSTEM_ROLE_ORDER = ["主设人", "设计人", "校核人", "审核人"]
PROJECT_STAFF_TEMPLATE_HEADERS = ["专业", "主设人", "设计人", "校核人", "审核人"]
# 人员名单分隔符：逗号/顿号/分号/换行/制表。刻意不把普通空格当分隔符，
# 以支持“张伟 线路”这类带空格的人名。
PROJECT_STAFF_SEPARATOR_PATTERN = re.compile(r"[、,，;；\r\n\t]+")

DISCIPLINE_ALIAS_TO_SYSTEM = {
    "设总": "项目设总",
    "项目设总": "项目设总",
    "结构": "土建结构",
    "土建结构": "土建结构",
    "建筑": "土建建筑",
    "土建建筑": "土建建筑",
    "资源": "资源",
    "技经": "技经",
    "综合能源": "综合能源",
    "电气一次": "电气一次",
    "电气二次": "电气二次",
    "线路电气": "线路电气",
    "线路结构": "线路结构",
    "总图": "总图",
    "道路": "道路",
    "水暖": "水暖",
    "岩土": "岩土",
    "测绘": "测绘",
}

# 允许的“人员专业”兼容关系：行专业（系统）可接受的人员专业（人员主数据中的原始专业）
SYSTEM_DEPT_ALLOWED_PERSON_DISCIPLINES = {
    "项目设总": {"项目设总", "设总", "设代"},
    "土建结构": {"土建结构", "结构"},
    "土建建筑": {"土建建筑", "建筑"},
    "综合能源": {"综合能源", "综合能源"},
}

# 系统专业 -> 固定日薪库“专业”候选（按顺序尝试）
SYSTEM_DEPT_TO_WAGE_DISCIPLINES = {
    "项目设总": ["设总", "设代"],
    "土建结构": ["结构"],
    "土建建筑": ["建筑"],
    "综合能源": ["综合能源"],
}


def normalize_system_discipline_name(val):
    name = str(val or "").strip()
    if not name:
        return ""
    if name in DISCIPLINE_ORDER:
        return name
    return DISCIPLINE_ALIAS_TO_SYSTEM.get(name, "")


def split_person_names(raw_text):
    text = str(raw_text or "").strip()
    if not text:
        return []
    items = [normalize_person_name(x) for x in PROJECT_STAFF_SEPARATOR_PATTERN.split(text) if x and str(x).strip()]
    ordered = []
    seen = set()
    for name in items:
        if name in seen:
            continue
        seen.add(name)
        ordered.append(name)
    return ordered


def dedupe_person_names(names):
    ordered = []
    seen = set()
    for raw in names or []:
        name = normalize_person_name(raw)
        if not name:
            continue
        if name in seen:
            continue
        seen.add(name)
        ordered.append(name)
    return ordered


def normalize_person_name(value):
    name = str(value or "").strip()
    if not name:
        return ""
    # 统一破折号/连字符、去掉连字符两侧空格
    name = name.replace("—", "-").replace("－", "-").replace("–", "-")
    name = re.sub(r"\s*-\s*", "-", name)
    # 连续空白归一
    name = re.sub(r"\s+", " ", name).strip()
    return name


def person_name_keys(value):
    base = normalize_person_name(value)
    if not base:
        return []
    keys = [base]
    no_space = base.replace(" ", "")
    if no_space and no_space not in keys:
        keys.append(no_space)
    return keys


def normalize_region_name(val):
    region = str(val or "").strip()
    if not region:
        return ""
    region = region.replace("所属区域：", "").replace("所属区域:", "").strip()
    return region


def person_matches_system_discipline(person_discipline, system_discipline):
    p_disc = str(person_discipline or "").strip()
    s_disc = str(system_discipline or "").strip()
    if not p_disc or not s_disc:
        return False
    allowed = SYSTEM_DEPT_ALLOWED_PERSON_DISCIPLINES.get(s_disc)
    if allowed:
        return p_disc in allowed
    normalized = normalize_system_discipline_name(p_disc)
    return normalized == s_disc


def get_wage_disciplines_for_system_dept(system_dept):
    dept = str(system_dept or "").strip()
    if not dept:
        return []
    candidates = list(SYSTEM_DEPT_TO_WAGE_DISCIPLINES.get(dept, []))
    if dept not in candidates:
        candidates.append(dept)
    return candidates


FIXED_WAGE_LOOKUP = {}
for rec in FIXED_WAGE_RECORDS:
    region = normalize_region_name(rec.get("region", ""))
    discipline = str(rec.get("discipline", "")).strip()
    grade = str(rec.get("grade", "")).strip()
    try:
        wage = float(rec.get("daily_wage", 0) or 0)
    except Exception:
        wage = 0.0
    if not region or not discipline or not grade or wage <= 0:
        continue
    FIXED_WAGE_LOOKUP[(region, discipline, grade)] = wage


FIXED_WAGE_GROUP_AVG = {}
_wage_group_acc = defaultdict(list)
for (region, discipline, grade), wage in FIXED_WAGE_LOOKUP.items():
    _wage_group_acc[(discipline, grade)].append(float(wage))
for key, vals in _wage_group_acc.items():
    if vals:
        FIXED_WAGE_GROUP_AVG[key] = sum(vals) / len(vals)


FIXED_WAGE_REGION_GRADE_AVG = {}
_wage_region_grade_acc = defaultdict(list)
for (region, discipline, grade), wage in FIXED_WAGE_LOOKUP.items():
    _wage_region_grade_acc[(region, grade)].append(float(wage))
for key, vals in _wage_region_grade_acc.items():
    if vals:
        FIXED_WAGE_REGION_GRADE_AVG[key] = sum(vals) / len(vals)


FIXED_WAGE_GRADE_AVG = {}
_wage_grade_acc = defaultdict(list)
for (region, discipline, grade), wage in FIXED_WAGE_LOOKUP.items():
    _wage_grade_acc[grade].append(float(wage))
for key, vals in _wage_grade_acc.items():
    if vals:
        FIXED_WAGE_GRADE_AVG[key] = sum(vals) / len(vals)


FIXED_PERSON_LOOKUP = defaultdict(list)
FIXED_PERSON_LOOKUP_NORM = defaultdict(list)
for person in FIXED_PERSON_MASTER:
    name = str(person.get("name", "")).strip()
    if not name:
        continue
    record = {
        "name": name,
        "discipline": str(person.get("discipline", "")).strip(),
        "region": normalize_region_name(person.get("region", "")),
        "grade": str(person.get("grade", "")).strip(),
    }
    FIXED_PERSON_LOOKUP[name].append(record)
    for key in person_name_keys(name):
        FIXED_PERSON_LOOKUP_NORM[key].append(record)

TASK_KEY_ALIASES = {
    "可行性研究（风电）": "可行性研究（风电项目，含升压站）",
    "可行性研究（光伏）": "可行性研究（光伏项目，含升压站）",
    "初步设计报告（风电）": "初步设计报告（风电项目，含升压站）",
    "初步设计报告（光伏）": "初步设计报告（光伏项目，含升压站）",
    "施工图（风电）": "施工图（风电项目，含升压站）",
    "施工图（光伏）": "施工图（光伏项目，含升压站）",
    "竣工图（风电）": "竣工图（风电项目，含升压站）",
    "竣工图（光伏）": "竣工图（光伏项目，含升压站）",
}

FEA_WITH_FIELD_BOOSTER = {
    "可行性研究（风电项目，含升压站）",
    "可行性研究（光伏项目，含升压站）",
}

FEA_STANDALONE_STATION = {
    "可行性研究（升压电站）",
    "可行性研究（储能电站）",
}

FEA_ALL_BOOSTER_RELATED = FEA_WITH_FIELD_BOOSTER | FEA_STANDALONE_STATION

PRELIM_ALL_BOOSTER_RELATED = {
    "初步设计报告（风电项目，含升压站）",
    "初步设计报告（光伏项目，含升压站）",
    "初步设计报告（升压电站）",
    "初步设计报告（储能电站）",
}

PRELIM_TO_FEA_TASK = {
    "初步设计报告（风电项目，含升压站）": "可行性研究（风电项目，含升压站）",
    "初步设计报告（光伏项目，含升压站）": "可行性研究（光伏项目，含升压站）",
    "初步设计报告（升压电站）": "可行性研究（升压电站）",
    "初步设计报告（储能电站）": "可行性研究（储能电站）",
}

PRELIM_WITH_OUTGOING = {
    "初步设计报告（升压电站）",
    "初步设计报告（储能电站）",
}

PRELIM_WITH_LINE_SURVEY = {
    "初步设计报告（风电项目，含升压站）",
    "初步设计报告（升压电站）",
    "初步设计报告（储能电站）",
}

PRELIM_MICRO_BONUS_RULES = {
    "初步设计报告（风电项目，含升压站）": {
        "related_fea": "可行性研究（风电项目，含升压站）",
        "days": {"资源": 5.0, "道路": 5.0},
    },
    "初步设计报告（升压电站）": {
        "related_fea": "可行性研究（升压电站）",
        "days": {"总图": 3.0},
    },
}

CONSTRUCTION_STATION_TASKS = {
    "施工图（升压电站）",
    "施工图（储能电站）",
}

OUTGOING_FEA_TASKS = {
    "可行性研究（升压电站）",
    "可行性研究（储能电站）",
}

OUTGOING_PRELIM_TASKS = PRELIM_WITH_OUTGOING
OUTGOING_CONSTRUCTION_TASKS = CONSTRUCTION_STATION_TASKS
OUTGOING_ELIGIBLE_TASKS = (
    OUTGOING_FEA_TASKS
    | OUTGOING_PRELIM_TASKS
    | OUTGOING_CONSTRUCTION_TASKS
)

STORAGE_BUNDLE_TASKS = {
    "可行性研究（储能电站）",
    "初步设计报告（储能电站）",
}


def clean_excel_header(val):
    return str(val or "").replace("\n", "").replace(" ", "").strip()


def canonical_task_key(task_key):
    key = str(task_key or "").strip()
    return TASK_KEY_ALIASES.get(key, key)


def get_task_config_search_paths():
    script_dir = os.path.abspath(os.path.dirname(__file__))
    script_path = os.path.join(script_dir, TASK_CONFIG_FILE)
    cwd_path = os.path.abspath(TASK_CONFIG_FILE)
    if cwd_path == script_path:
        return [script_path]
    return [script_path, cwd_path]


def get_task_config_save_path():
    return get_task_config_search_paths()[0]


def infer_task_targets(task_key):
    key = canonical_task_key(task_key)
    if "（风电" in key:
        return ("wind",)
    if "（光伏" in key:
        return ("pv",)
    if "（升压电站）" in key or "（储能电站）" in key:
        return ("wind", "pv")
    return ("wind", "pv")


def normalize_existing_task_keys():
    # 将旧键名统一到新键名，避免历史配置与新版任务名不一致。
    for item in DATA_WIND_PARAMS:
        old_key = item["key"]
        new_key = canonical_task_key(old_key)
        if new_key == old_key:
            continue
        item["key"] = new_key
        if old_key in RAW_MATRIX_WIND and new_key not in RAW_MATRIX_WIND:
            RAW_MATRIX_WIND[new_key] = RAW_MATRIX_WIND.pop(old_key)

    for item in DATA_PV_PARAMS:
        old_key = item["key"]
        new_key = canonical_task_key(old_key)
        if new_key == old_key:
            continue
        item["key"] = new_key
        if old_key in RAW_MATRIX_PV and new_key not in RAW_MATRIX_PV:
            RAW_MATRIX_PV[new_key] = RAW_MATRIX_PV.pop(old_key)


def normalize_excel_task_name(project_type, detail_text):
    key = str(project_type or "").strip()
    detail_text = str(detail_text or "")
    if not key:
        return ""

    if key == "投标技术文件":
        detail_upper = detail_text.upper()
        if "EPC" in detail_upper:
            return "投标技术文件 (EPC)"
        if "咨询" in detail_text or "咨" in detail_text:
            return "投标技术文件 (咨询)"
        if "勘察" in detail_text or "勘" in detail_text:
            return "投标技术文件 (勘察)"

    return canonical_task_key(key)


def is_red_rich_segment(part):
    font = getattr(part, "font", None)
    if not font:
        return False
    color = getattr(font, "color", None)
    if color is None:
        return False
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str):
        return rgb.upper().endswith("FF0000")
    return False


def parse_detail_segments(cell_value):
    if cell_value is None:
        return []
    if type(cell_value).__name__ == "CellRichText":
        segments = []
        for part in cell_value:
            txt = str(part)
            if not txt:
                continue
            segments.append({"text": txt, "red": is_red_rich_segment(part)})
        return segments
    txt = str(cell_value)
    return [{"text": txt, "red": False}] if txt else []


def normalize_intl_flag(raw_value, default=None):
    if raw_value is None:
        return default
    if isinstance(raw_value, bool):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return abs(float(raw_value)) > 1e-12

    text = str(raw_value).strip()
    if not text:
        return default

    lowered = text.lower().replace(" ", "")
    false_tokens = {"-", "0", "n", "no", "false", "否", "无", "不适用"}
    true_tokens = {
        "1",
        "y",
        "yes",
        "true",
        "是",
        "适用",
        "国际",
        "国际标准",
        "√",
        "✓",
        "✔",
        "✔️",
    }

    if lowered in false_tokens:
        return False
    if lowered in true_tokens:
        return True

    if "否" in text or "不适用" in text:
        return False
    if "是" in text or "国际" in text or "适用" in text:
        return True

    return default


def parse_task_config_excel(file_path):
    if not HAS_OPENPYXL:
        raise RuntimeError("未安装 openpyxl，无法读取 Excel。")

    try:
        wb = load_workbook(file_path, data_only=False, rich_text=True)
    except TypeError:
        wb = load_workbook(file_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    for r in range(1, min(12, ws.max_row + 1)):
        row_headers = [clean_excel_header(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "项目类型" in row_headers and "项目总工日" in row_headers:
            header_row = r
            break
    if header_row is None:
        raise ValueError("未识别到 Excel 表头（项目类型/项目总工日）。")

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    col_project_type = None
    col_detail = None
    col_total_days = None
    col_intl = None
    ratio_cols = {}

    for c, h in enumerate(headers, start=1):
        h_clean = clean_excel_header(h)
        if h_clean == "项目类型":
            col_project_type = c
        elif h_clean == "详情":
            col_detail = c
        elif h_clean == "项目总工日":
            col_total_days = c
        elif ("国际" in h_clean and "系数" in h_clean) or ("intl" in h_clean.lower()):
            col_intl = c
        elif h_clean in RATIO_COLUMN_MAP:
            ratio_cols[c] = RATIO_COLUMN_MAP[h_clean]

    if col_project_type is None or col_total_days is None:
        raise ValueError("Excel 缺少关键列：项目类型 / 项目总工日。")
    if not ratio_cols:
        raise ValueError("Excel 未识别到专业工日分配列。")

    task_rows = {}
    task_details = {}
    for r in range(header_row + 1, ws.max_row + 1):
        project_type = ws.cell(r, col_project_type).value
        total_days_raw = ws.cell(r, col_total_days).value
        if not project_type:
            continue

        if isinstance(total_days_raw, (int, float)):
            total_days = float(total_days_raw)
        else:
            total_days = 0.0
            for c in ratio_cols.keys():
                raw_val = ws.cell(r, c).value
                if isinstance(raw_val, (int, float)):
                    total_days += float(raw_val)

        if total_days <= 0:
            continue

        detail_val = ws.cell(r, col_detail).value if col_detail else ""
        task_key = normalize_excel_task_name(project_type, detail_val)
        if not task_key:
            continue

        ratios = {}
        for c, dept in ratio_cols.items():
            raw_val = ws.cell(r, c).value
            if not isinstance(raw_val, (int, float)):
                continue
            raw_val = float(raw_val)
            if abs(raw_val) < 1e-12:
                continue
            ratios[dept] = round(raw_val * 100.0 / total_days, 2)

        row_cfg = {"base_days": total_days, "ratios": ratios}
        if col_intl is not None:
            # In current Excel templates, "-" means not applicable; blank is treated as applicable.
            intl_flag = normalize_intl_flag(ws.cell(r, col_intl).value, default=True)
            row_cfg["intl"] = bool(intl_flag)

        task_rows[task_key] = row_cfg
        detail_segments = parse_detail_segments(detail_val)
        if detail_segments:
            task_details[task_key] = detail_segments

    return {"tasks": task_rows, "details": task_details}


def apply_task_config_overrides(payload, persist=False, source_file=None, replace_all_tasks=False):
    global MATRIX_WIND, MATRIX_PV, DB_WIND_PARAMS, DB_PV_PARAMS, CURRENT_TASK_DETAILS
    if not isinstance(payload, dict):
        return {"updated_wind": 0, "updated_pv": 0}

    normalize_existing_task_keys()

    tasks_in = payload.get("tasks", {})
    details_in = payload.get("details", {})
    tasks = {}
    details = {}

    if isinstance(tasks_in, dict):
        for key, cfg in tasks_in.items():
            ckey = canonical_task_key(key)
            if isinstance(cfg, dict):
                tasks[ckey] = cfg

    if isinstance(details_in, dict):
        for key, segs in details_in.items():
            ckey = canonical_task_key(key)
            if isinstance(segs, list):
                details[ckey] = segs

    # Optional replace mode: prune tasks/matrix not present in payload.
    # This is useful when latest Excel intentionally removes tasks.
    if replace_all_tasks and tasks:
        keep_keys = set(tasks.keys())
        DATA_WIND_PARAMS[:] = [item for item in DATA_WIND_PARAMS if canonical_task_key(item.get("key")) in keep_keys]
        DATA_PV_PARAMS[:] = [item for item in DATA_PV_PARAMS if canonical_task_key(item.get("key")) in keep_keys]

        for old_key in list(RAW_MATRIX_WIND.keys()):
            if canonical_task_key(old_key) not in keep_keys:
                RAW_MATRIX_WIND.pop(old_key, None)
        for old_key in list(RAW_MATRIX_PV.keys()):
            if canonical_task_key(old_key) not in keep_keys:
                RAW_MATRIX_PV.pop(old_key, None)

        if CURRENT_TASK_DETAILS:
            CURRENT_TASK_DETAILS = {
                canonical_task_key(k): v
                for k, v in CURRENT_TASK_DETAILS.items()
                if canonical_task_key(k) in keep_keys
            }

    wind_keys = {item["key"] for item in DATA_WIND_PARAMS}
    pv_keys = {item["key"] for item in DATA_PV_PARAMS}

    for key, cfg in tasks.items():
        targets = infer_task_targets(key)
        base_days = float(cfg.get("base_days", 0) or 0)
        ratios = cfg.get("ratios", {})
        intl_value = normalize_intl_flag(cfg.get("intl", None), default=None)
        if "wind" in targets:
            if key not in wind_keys:
                DATA_WIND_PARAMS.append(
                    {
                        "key": key,
                        "base_days": base_days,
                        "rule": None,
                        "intl": bool(intl_value) if intl_value is not None else True,
                    }
                )
                wind_keys.add(key)
            if isinstance(ratios, dict) and ratios:
                RAW_MATRIX_WIND[key] = {k: float(v) for k, v in ratios.items()}
        if "pv" in targets:
            if key not in pv_keys:
                DATA_PV_PARAMS.append(
                    {
                        "key": key,
                        "base_days": base_days,
                        "rule": None,
                        "intl": bool(intl_value) if intl_value is not None else True,
                    }
                )
                pv_keys.add(key)
            if isinstance(ratios, dict) and ratios:
                RAW_MATRIX_PV[key] = {k: float(v) for k, v in ratios.items()}

    updated_wind = 0
    updated_pv = 0

    for item in DATA_WIND_PARAMS:
        key = item["key"]
        cfg = tasks.get(key)
        if not cfg:
            continue
        item["base_days"] = float(cfg.get("base_days", item["base_days"]))
        intl_value = normalize_intl_flag(cfg.get("intl", None), default=None)
        if intl_value is not None:
            item["intl"] = bool(intl_value)
        ratios = cfg.get("ratios", {})
        if isinstance(ratios, dict) and ratios:
            RAW_MATRIX_WIND[key] = {k: float(v) for k, v in ratios.items()}
        updated_wind += 1

    for item in DATA_PV_PARAMS:
        key = item["key"]
        cfg = tasks.get(key)
        if not cfg:
            continue
        item["base_days"] = float(cfg.get("base_days", item["base_days"]))
        intl_value = normalize_intl_flag(cfg.get("intl", None), default=None)
        if intl_value is not None:
            item["intl"] = bool(intl_value)
        ratios = cfg.get("ratios", {})
        if isinstance(ratios, dict) and ratios:
            RAW_MATRIX_PV[key] = {k: float(v) for k, v in ratios.items()}
        updated_pv += 1

    MATRIX_WIND = normalize_matrix(RAW_MATRIX_WIND)
    MATRIX_PV = normalize_matrix(RAW_MATRIX_PV)
    DB_WIND_PARAMS = {x["key"]: {**x, "parsed_rule": parse_rules(x["rule"])} for x in DATA_WIND_PARAMS}
    DB_PV_PARAMS = {x["key"]: {**x, "parsed_rule": parse_rules(x["rule"])} for x in DATA_PV_PARAMS}

    if details:
        normalized_details = {}
        for key, segs in CURRENT_TASK_DETAILS.items():
            normalized_details[canonical_task_key(key)] = segs
        for key, segs in details.items():
            normalized_details[key] = segs
        CURRENT_TASK_DETAILS = normalized_details

    if persist:
        save_path = get_task_config_save_path()
        save_json(
            save_path,
            {
                "source_file": source_file or "",
                "updated_at": datetime.datetime.now().isoformat(timespec="seconds"),
                "tasks": tasks,
                "details": CURRENT_TASK_DETAILS,
            },
        )

    return {"updated_wind": updated_wind, "updated_pv": updated_pv}


def load_task_config_overrides_from_file():
    global MATRIX_WIND, MATRIX_PV, DB_WIND_PARAMS, DB_PV_PARAMS
    normalize_existing_task_keys()
    MATRIX_WIND = normalize_matrix(RAW_MATRIX_WIND)
    MATRIX_PV = normalize_matrix(RAW_MATRIX_PV)
    DB_WIND_PARAMS = {x["key"]: {**x, "parsed_rule": parse_rules(x["rule"])} for x in DATA_WIND_PARAMS}
    DB_PV_PARAMS = {x["key"]: {**x, "parsed_rule": parse_rules(x["rule"])} for x in DATA_PV_PARAMS}
    for path in get_task_config_search_paths():
        if not os.path.exists(path):
            continue
        payload = load_json(path, {})
        if not isinstance(payload, dict):
            continue
        if "tasks" not in payload and "details" not in payload:
            continue
        # Keep built-in tasks and only override those provided by Excel.
        # This avoids accidentally dropping legacy tasks (e.g. 竣工图) when
        # the latest Excel omits them.
        apply_task_config_overrides(payload, persist=False, replace_all_tasks=False)
        return


load_task_config_overrides_from_file()

def get_dept_sort_index(dept_name):
    if dept_name in DISCIPLINE_ORDER: return DISCIPLINE_ORDER.index(dept_name)
    for i, d in enumerate(DISCIPLINE_ORDER):
        if d in dept_name: return i
    return 999

def get_fixed_wage(dept_name, role_name):
    if dept_name in WAGE_DATABASE:
        return WAGE_DATABASE[dept_name].get(role_name, 0)
    for key, wages in WAGE_DATABASE.items():
        if key in dept_name:
            return wages.get(role_name, 0)
    return WAGE_DATABASE["DEFAULT"].get(role_name, 0)

# ==========================================
# UI 主题与样式
# ==========================================
UI_FONT = ("Microsoft YaHei UI", 9)
UI_FONT_BOLD = ("Microsoft YaHei UI", 9, "bold")
UI_FONT_SMALL = ("Microsoft YaHei UI", 8)
UI_FONT_TITLE = ("Microsoft YaHei UI", 16, "bold")

UI_COLORS = {
    "bg": "#F5F7FB",
    "card": "#FFFFFF",
    "text": "#1B2430",
    "muted": "#667085",
    "border": "#D6DFEA",
    "primary": "#2F6BFF",
    "primary_hover": "#265BE0",
    "primary_active": "#1F4CC4",
    "danger": "#E11D48",
    "warning_bg": "#FFF7ED",
}

HELP_TEXT = (
    "使用说明：\n"
    "1. 选择项目类型（风电/光伏/储能）并填写容量(MW)\n"
    "2. 导入项目人员名单（专业+主设/设计/校核/审核）\n"
    "3. 选择地形与国际系数（如适用）\n"
    "4. 点击“生成人工预算表”生成明细\n"
    "5. “导出精美报表”将生成带封面、汇总与明细的 xlsx\n"
    "6. 可导出/导入模板（xlsx 或 csv）\n"
    "\n"
    "创作人：张钰、王雨\n"
    "联系电话：张钰 13521243226，王雨 18810950371\n"
)

def apply_theme(root):
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except:
        pass

    root.configure(bg=UI_COLORS["bg"])
    root.option_add("*Font", UI_FONT)

    style.configure(".", font=UI_FONT, foreground=UI_COLORS["text"])
    style.configure("TFrame", background=UI_COLORS["bg"])
    style.configure("TLabelframe", background=UI_COLORS["bg"], bordercolor=UI_COLORS["border"])
    style.configure("TLabelframe.Label", background=UI_COLORS["bg"], foreground=UI_COLORS["text"], font=UI_FONT_BOLD)

    style.configure("TLabel", background=UI_COLORS["bg"], foreground=UI_COLORS["text"])
    style.configure("Field.TLabel", foreground=UI_COLORS["muted"], font=UI_FONT_SMALL, background=UI_COLORS["bg"])
    style.configure("Hint.TLabel", foreground=UI_COLORS["muted"], font=UI_FONT_SMALL, background=UI_COLORS["bg"])
    style.configure("Cost.TLabel", foreground=UI_COLORS["primary"], font=UI_FONT_BOLD, background=UI_COLORS["bg"])
    style.configure("Summary.TLabel", foreground=UI_COLORS["danger"], font=UI_FONT_BOLD, background=UI_COLORS["warning_bg"], padding=(10, 8))

    style.configure("TButton", padding=(12, 6))
    style.map("TButton", background=[("active", "#E9EEF7"), ("pressed", "#D9E2F1")])
    style.configure("Primary.TButton", background=UI_COLORS["primary"], foreground="#FFFFFF", padding=(12, 6), font=UI_FONT_BOLD)
    style.map(
        "Primary.TButton",
        background=[("active", UI_COLORS["primary_hover"]), ("pressed", UI_COLORS["primary_active"])],
        foreground=[("disabled", "#E5E7EB")],
    )

    style.configure("TEntry", padding=(6, 4))
    style.configure("TCombobox", padding=(6, 4))
    style.configure("TCheckbutton", background=UI_COLORS["bg"])
    style.configure("TRadiobutton", background=UI_COLORS["bg"])
    style.configure("TSeparator", background=UI_COLORS["border"])

# ==========================================
# 3. 基础UI组件
# ==========================================
class LabelEntry(ttk.Frame):
    def __init__(self, parent, label, value="", width=8, readonly=False, callback=None):
        super().__init__(parent)
        # REMOVED self.pack() to allow grid layout
        
        self.label_widget = ttk.Label(self, text=label, style="Field.TLabel")
        self.label_widget.pack(anchor="w")
        state = "readonly" if readonly else "normal"
        self.var = tk.StringVar(value=str(value))
        self.entry = ttk.Entry(self, textvariable=self.var, width=width, state=state)
        self.entry.pack(anchor="w")
        if callback:
            self.entry.bind("<Return>", lambda e: callback())
            self.entry.bind("<FocusOut>", lambda e: callback())
     
    def get(self): return self.var.get()
    def set(self, val): self.var.set(val)
    def set_label(self, text): self.label_widget.configure(text=text)

class SearchableCombobox(ttk.Combobox):
    def __init__(self, master=None, all_values=None, callback_select=None, **kwargs):
        self.var = tk.StringVar()
        super().__init__(master, textvariable=self.var, **kwargs)
        self.all_values = sorted(all_values) if all_values else []
        self.callback_select = callback_select
        self['values'] = self.all_values
        self.var.trace('w', self.on_text_change)
        self.bind('<<ComboboxSelected>>', self.on_select)
        self.bind('<Button-1>', self.on_click)

    def set_all_values(self, values):
        self.all_values = sorted(values)
        self['values'] = self.all_values

    def set_value(self, value):
        self.set(value)
        if self.callback_select: self.callback_select()

    def on_click(self, event):
        if not self.get().strip(): self['values'] = self.all_values

    def on_text_change(self, *args):
        text = self.var.get().strip()
        if not text: self['values'] = self.all_values
        else:
            filtered = [v for v in self.all_values if text.lower() in v.lower()]
            self['values'] = filtered
            if filtered:
                try: self.event_generate('<Down>')
                except: pass

    def on_select(self, event):
        if self.callback_select: self.callback_select()

class ScrollableCheckBoxFrame(ttk.Frame):
    def __init__(self, container, height=220, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(self, height=height, highlightthickness=0, bg=UI_COLORS["bg"])
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        self.canvas.bind(
            "<Configure>",
            lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width)
        )

        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        self.bind_scroll(self.canvas)
        self.bind_scroll(self.scrollable_frame)
        self.vars = {} 

    def bind_scroll(self, widget):
        widget.bind("<MouseWheel>", self.on_mousewheel)
        widget.bind("<Button-4>", self.on_mousewheel)
        widget.bind("<Button-5>", self.on_mousewheel)

    def on_mousewheel(self, event):
        if platform.system() == "Windows":
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        else:
            if event.num == 4: self.canvas.yview_scroll(-1, "units")
            elif event.num == 5: self.canvas.yview_scroll(1, "units")

    def populate(self, items):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.vars = {}
        for item in items:
            var = tk.BooleanVar(value=False)
            row = ttk.Frame(self.scrollable_frame)
            row.pack(fill="x", padx=5, pady=2)
            chk = ttk.Checkbutton(row, text=item, variable=var)
            chk.pack(anchor="w")
            self.bind_scroll(chk)
            self.vars[item] = var
        self.scrollable_frame.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def get_checked_items(self):
        return [item for item, var in self.vars.items() if var.get()]
    def select_all(self):
        for var in self.vars.values(): var.set(True)
    def deselect_all(self):
        for var in self.vars.values(): var.set(False)


class TaskSelectionDialog(tk.Toplevel):
    def __init__(self, parent, task_items, preselected=None, on_confirm=None):
        super().__init__(parent)
        self.title("任务选择")
        self.geometry("680x620")
        self.minsize(560, 460)
        self.configure(bg=UI_COLORS["bg"])
        self.on_confirm = on_confirm

        ttk.Label(self, text="请选择任务（可多选），然后点击“确认选择”", style="Hint.TLabel").pack(
            anchor="w", padx=10, pady=(10, 4)
        )

        tools = ttk.Frame(self)
        tools.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Button(tools, text="全选", command=self._select_all).pack(side="left")
        ttk.Button(tools, text="清空", command=self._deselect_all).pack(side="left", padx=6)
        self.lbl_count = ttk.Label(tools, text="已选: 0 项", style="Hint.TLabel")
        self.lbl_count.pack(side="right")

        self.chk_frame = ScrollableCheckBoxFrame(self, height=480)
        self.chk_frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        self.chk_frame.populate(task_items or [])

        selected_set = set(preselected or [])
        for task, var in self.chk_frame.vars.items():
            var.set(task in selected_set)
            var.trace_add("write", lambda *_: self._update_count())
        self._update_count()

        f_btn = ttk.Frame(self)
        f_btn.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(f_btn, text="确认选择", style="Primary.TButton", command=self._confirm).pack(side="right")
        ttk.Button(f_btn, text="取消", command=self.destroy).pack(side="right", padx=(0, 8))

        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _update_count(self):
        if hasattr(self, "lbl_count"):
            self.lbl_count.configure(text=f"已选: {len(self.chk_frame.get_checked_items())} 项")

    def _select_all(self):
        self.chk_frame.select_all()
        self._update_count()

    def _deselect_all(self):
        self.chk_frame.deselect_all()
        self._update_count()

    def _confirm(self):
        if self.on_confirm:
            self.on_confirm(self.chk_frame.get_checked_items())
        self.destroy()


class ExtrasSelectionDialog(tk.Toplevel):
    def __init__(self, parent, state, project_type="风电", on_confirm=None):
        super().__init__(parent)
        self.title("其他修订项")
        self.geometry("760x520")
        self.minsize(640, 420)
        self.configure(bg=UI_COLORS["bg"])
        self.on_confirm = on_confirm
        self.project_type = project_type

        state = state or {}
        self.v_ext_micro = tk.BooleanVar(value=bool(state.get("ext_micro", False)))
        self.v_ext_turbine = tk.BooleanVar(value=bool(state.get("ext_turbine", False)))
        self.v_ext_proposal = tk.BooleanVar(value=bool(state.get("ext_proposal", False)))
        self.v_ext_app = tk.BooleanVar(value=bool(state.get("ext_app", False)))
        self.v_ext_bid = tk.BooleanVar(value=bool(state.get("ext_bid", False)))
        self.v_grid_review = tk.BooleanVar(value=bool(state.get("grid_review", False)))
        self.v_has_outgoing = tk.BooleanVar(value=bool(state.get("has_outgoing", False)))
        self.v_has_line_survey = tk.BooleanVar(value=bool(state.get("has_line_survey", False)))
        self.v_storage_bundle = tk.BooleanVar(value=bool(state.get("storage_bundle", False)))
        self.v_prelim_micro_bonus = tk.BooleanVar(value=bool(state.get("prelim_micro_bonus", False)))
        self.v_line_survey_terrain = tk.StringVar(value=state.get("line_survey_terrain", "非山地"))
        self.v_booster_kv = tk.StringVar(value=state.get("booster_kv", "常规"))

        ttk.Label(self, text="请选择修订项，确认后生效", style="Hint.TLabel").pack(anchor="w", padx=10, pady=(10, 6))

        area = SectionScrollArea(self, height=410, fit_width=True, with_hscroll=False)
        area.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        body = area.body

        f_ex1 = ttk.Frame(body); f_ex1.pack(fill="x", anchor="w", padx=4, pady=(1, 0))
        ttk.Checkbutton(f_ex1, text="微观选址(+6)", variable=self.v_ext_micro).pack(side="left", padx=(0, 10))
        ttk.Checkbutton(f_ex1, text="机型比选(+10)", variable=self.v_ext_turbine).pack(side="left")

        f_ex2 = ttk.Frame(body); f_ex2.pack(fill="x", anchor="w", padx=4, pady=(1, 0))
        ttk.Checkbutton(f_ex2, text="建议书(+3)", variable=self.v_ext_proposal).pack(side="left", padx=(0, 10))
        ttk.Checkbutton(f_ex2, text="项目申请(+6)", variable=self.v_ext_app).pack(side="left", padx=(0, 10))
        ttk.Checkbutton(f_ex2, text="竞配报告(+3)", variable=self.v_ext_bid).pack(side="left")

        f_ex3 = ttk.Frame(body); f_ex3.pack(fill="x", anchor="w", padx=4, pady=(1, 0))
        ttk.Checkbutton(f_ex3, text="配合电网评审", variable=self.v_grid_review).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(
            f_ex3,
            text="含外送（仅适用于单独的升压站/储能站的可研、初设、施工图）",
            variable=self.v_has_outgoing,
        ).pack(side="left")

        f_ex4 = ttk.Frame(body); f_ex4.pack(fill="x", anchor="w", padx=4, pady=(1, 0))
        ttk.Checkbutton(f_ex4, text="含线路测绘（仅初设）", variable=self.v_has_line_survey).pack(side="left")
        ttk.Label(f_ex4, text="地形:").pack(side="left", padx=(10, 0))
        self.c_line_survey_terrain = ttk.Combobox(
            f_ex4, width=8, state="readonly", values=["非山地", "山地"], textvariable=self.v_line_survey_terrain
        )
        if self.v_line_survey_terrain.get() not in ("非山地", "山地"):
            self.v_line_survey_terrain.set("非山地")
        self.c_line_survey_terrain.pack(side="left", padx=(4, 0))

        f_ex5 = ttk.Frame(body); f_ex5.pack(fill="x", anchor="w", padx=4, pady=(1, 2))
        ttk.Label(f_ex5, text="升压站电压等级:").pack(side="left")
        self.c_booster_kv = ttk.Combobox(
            f_ex5, width=8, state="readonly", values=["常规", "220kV", "330kV"], textvariable=self.v_booster_kv
        )
        if self.v_booster_kv.get() not in ("常规", "220kV", "330kV"):
            self.v_booster_kv.set("常规")
        self.c_booster_kv.pack(side="left", padx=(4, 12))
        self.chk_storage_bundle = ttk.Checkbutton(
            f_ex5,
            text="若为配储(综合能源+8，仅储能项目适用)",
            variable=self.v_storage_bundle,
        )
        self.chk_storage_bundle.pack(side="left")
        if self.project_type != "储能":
            self.v_storage_bundle.set(False)
            self.chk_storage_bundle.state(["disabled"])

        f_ex6 = ttk.Frame(body); f_ex6.pack(fill="x", anchor="w", padx=4, pady=(1, 2))
        ttk.Checkbutton(
            f_ex6,
            text="初设微选附加（风电:资源+5/道路+5；升压站:总图+3）",
            variable=self.v_prelim_micro_bonus,
        ).pack(side="left")

        f_btn = ttk.Frame(self)
        f_btn.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(f_btn, text="确认选择", style="Primary.TButton", command=self._confirm).pack(side="right")
        ttk.Button(f_btn, text="取消", command=self.destroy).pack(side="right", padx=(0, 8))

        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _confirm(self):
        if self.on_confirm:
            self.on_confirm(
                {
                    "ext_micro": self.v_ext_micro.get(),
                    "ext_turbine": self.v_ext_turbine.get(),
                    "ext_proposal": self.v_ext_proposal.get(),
                    "ext_app": self.v_ext_app.get(),
                    "ext_bid": self.v_ext_bid.get(),
                    "grid_review": self.v_grid_review.get(),
                    "has_outgoing": self.v_has_outgoing.get(),
                    "has_line_survey": self.v_has_line_survey.get(),
                    "line_survey_terrain": self.v_line_survey_terrain.get().strip() or "非山地",
                    "booster_kv": self.v_booster_kv.get().strip() or "常规",
                    "storage_bundle": self.v_storage_bundle.get(),
                    "prelim_micro_bonus": self.v_prelim_micro_bonus.get(),
                }
            )
        self.destroy()


class SectionScrollArea(ttk.Frame):
    def __init__(self, container, height=120, fit_width=True, with_hscroll=False, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.fit_width = fit_width
        self.canvas = tk.Canvas(self, height=height, highlightthickness=0, bg=UI_COLORS["bg"])
        self.v_scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.v_scrollbar.set)

        self.h_scrollbar = None
        if with_hscroll:
            self.h_scrollbar = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
            self.canvas.configure(xscrollcommand=self.h_scrollbar.set)

        if self.h_scrollbar:
            self.h_scrollbar.pack(side="bottom", fill="x")
        self.v_scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.body = ttk.Frame(self.canvas)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.body, anchor="nw")
        self.body.bind("<Configure>", self._on_body_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self._bind_scroll(self.canvas)
        self._bind_scroll(self.body)

    def _on_body_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        if self.fit_width:
            self.canvas.itemconfigure(self.canvas_window, width=event.width)
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _bind_scroll(self, widget):
        widget.bind("<MouseWheel>", self._on_mousewheel)
        widget.bind("<Button-4>", self._on_mousewheel)
        widget.bind("<Button-5>", self._on_mousewheel)

    def _on_mousewheel(self, event):
        if platform.system() == "Windows":
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        else:
            if event.num == 4:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.canvas.yview_scroll(1, "units")

class GeotechProcurementPanel(ttk.LabelFrame):
    def __init__(self, parent, update_callback):
        super().__init__(parent, text="4. 地勘采购成本", padding=5)
        self.update_callback = update_callback
        self.current_project_type = "风电"
        self.init_ui()
        
    def init_ui(self):
        self.v_enable = tk.BooleanVar(value=True)
        geo_header = ttk.Frame(self)
        ttk.Label(geo_header, text="4. 地勘采购成本").pack(side="left")
        ttk.Checkbutton(
            geo_header,
            text="计入总造价",
            variable=self.v_enable,
            command=self.on_manual_calc,
        ).pack(side="left", padx=(8, 0))
        self.configure(text="")
        self.configure(labelwidget=geo_header)

        self.geo_scroll = SectionScrollArea(self, height=170, fit_width=True, with_hscroll=True)
        self.geo_scroll.pack(fill="both", expand=True)
        body = self.geo_scroll.body

        f1 = ttk.Frame(body)
        f1.pack(fill="x", pady=2)
        ttk.Label(f1, text="阶段:").grid(row=0, column=0, sticky="w")
        self.c_stage = ttk.Combobox(f1, values=["初勘", "详勘"], width=5, state="readonly")
        self.c_stage.current(0)
        self.c_stage.grid(row=0, column=1, padx=2)
        self.c_stage.bind("<<ComboboxSelected>>", self.on_auto_calc)

        ttk.Label(f1, text="地质:").grid(row=0, column=2, padx=(5,0), sticky="w")
        self.c_soil = ttk.Combobox(f1, values=["土层", "碎石土", "岩石"], width=6, state="readonly")
        self.c_soil.current(0)
        self.c_soil.grid(row=0, column=3, padx=2)
        self.c_soil.bind("<<ComboboxSelected>>", self.on_auto_calc) 
        
        self.e_scale = LabelEntry(f1, "风机台数", "8", width=8, callback=self.on_auto_calc)
        # LabelEntry 没有 self.pack() 了，所以这里手动 grid
        self.e_scale.grid(row=0, column=4, padx=5)
        
        f2 = ttk.Frame(body)
        f2.pack(fill="x", pady=2)
        self.v_booster = tk.BooleanVar(value=True)
        self.v_storage = tk.BooleanVar(value=False)
        self.v_slope = tk.BooleanVar(value=False)
        ttk.Checkbutton(f2, text="含升压站", variable=self.v_booster, command=self.on_auto_calc).grid(row=0, column=0, sticky="w")
        ttk.Checkbutton(f2, text="含储能", variable=self.v_storage, command=self.on_auto_calc).grid(row=0, column=1, padx=5, sticky="w")
        ttk.Checkbutton(f2, text="含边坡", variable=self.v_slope, command=self.on_auto_calc).grid(row=0, column=2, padx=5, sticky="w")
        
        f3 = ttk.Frame(body)
        f3.pack(fill="x", pady=2)
        self.e_holes = LabelEntry(f3, "场区孔数", "0", width=4, readonly=False, callback=self.on_manual_calc)
        self.e_holes.pack(side="left", padx=2) # 手动 pack
        self.e_field_footage = LabelEntry(f3, "场区进尺", "0", width=6, readonly=False, callback=self.on_manual_calc)
        self.e_field_footage.pack(side="left", padx=2) # 手动 pack
        self.e_booster_footage = LabelEntry(f3, "升压站进尺", "0", width=6, readonly=False, callback=self.on_manual_calc)
        self.e_booster_footage.pack(side="left", padx=2) # 手动 pack
        
        f4 = ttk.Frame(body)
        f4.pack(fill="x", pady=2)
        self.e_price = LabelEntry(f4, "单价", "0", width=4, readonly=False, callback=self.on_manual_calc)
        self.e_price.pack(side="left", padx=2) # 手动 pack
        self.e_lab_fee = LabelEntry(f4, "实验费", "0", width=6, readonly=False, callback=self.on_manual_calc)
        self.e_lab_fee.pack(side="left", padx=2) # 手动 pack
        self.e_extra_fee = LabelEntry(f4, "杂费", "0", width=6, readonly=False, callback=self.on_manual_calc)
        self.e_extra_fee.pack(side="left", padx=2) # 手动 pack
        
        self.lbl_cost = ttk.Label(body, text="地勘费用: ¥0", style="Cost.TLabel")
        self.lbl_cost.pack(anchor="e", pady=5)
        
    def set_project_type(self, p_type):
        self.current_project_type = p_type
        if p_type in ("光伏", "储能"):
            self.e_scale.set_label("占地面积(亩)")
            self.e_scale.var.set("2000")
        else:
            self.e_scale.set_label("风机台数")
            self.e_scale.var.set("8")
        self.on_auto_calc()

    def on_auto_calc(self, event=None):
        p_type = self.current_project_type
        stage = self.c_stage.get()
        soil_type = self.c_soil.get()
        
        try: scale = float(self.e_scale.get())
        except: scale = 0 
        
        type_defaults = PROCUREMENT_DEFAULTS.get(p_type)
        if not type_defaults: return 

        holes = 0
        if p_type == "风电":
            if stage == "初勘":
                holes = max(math.ceil(scale / 10), 2)
            else:
                holes = int(scale)
        else:
            area_sqm = scale * 666.7 
            if area_sqm > 0:
                spacing = 300 if stage == "初勘" else 150
                side_nodes = math.sqrt(area_sqm) / spacing + 1
                holes = math.ceil(side_nodes ** 2)
            else:
                holes = 0
        
        self.e_holes.set(holes)
        
        if p_type == "风电":
            if soil_type == "土层": unit_depth = 50
            elif soil_type == "碎石土": unit_depth = 20
            else: unit_depth = 15 
        else:
            unit_depth = type_defaults["孔深"]
            
        field_footage = holes * unit_depth
        self.e_field_footage.set(field_footage)
        
        booster_footage = 0
        if self.v_booster.get():
            has_storage = self.v_storage.get()
            has_slope = self.v_slope.get()
            base_val = type_defaults["升压站基准"].get(stage, 0)
            adder = 0
            if stage == "初勘":
                if has_storage and has_slope: adder = 40
                elif has_storage or has_slope: adder = 20
            else: 
                if has_storage and has_slope: adder = 200
                elif has_storage or has_slope: adder = 100
            booster_footage = base_val + adder
                
        self.e_booster_footage.set(booster_footage)
        
        price = 0
        if p_type == "风电":
            if soil_type == "岩石": price = 220
            elif soil_type == "碎石土": price = 160
            else: price = 130 
        else:
            price = type_defaults["单价"].get(stage, 0)

        lab_fee = 0
        if p_type == "风电":
            if stage == "初勘":
                if holes <= 20: lab_fee = 10000
                elif holes <= 50: lab_fee = 15000
                else: lab_fee = 20000
            else:
                lab_fee = 10000
        else:
            if holes <= 50: lab_fee = 5000
            elif holes <= 100: lab_fee = 10000
            else: lab_fee = 15000

        extra_fee = 15000 if stage == "初勘" else 10000
        
        self.e_price.set(price)
        self.e_lab_fee.set(lab_fee)
        self.e_extra_fee.set(extra_fee)
        self.on_manual_calc()

    def on_manual_calc(self, event=None):
        try:
            field_f = float(self.e_field_footage.get())
            booster_f = float(self.e_booster_footage.get())
            total_footage = field_f + booster_f
            price = float(self.e_price.get())
            lab = float(self.e_lab_fee.get())
            extra = float(self.e_extra_fee.get())
            raw_cost = (total_footage * price) + lab + extra
            
            if self.v_enable.get():
                self.current_cost = raw_cost
                self.lbl_cost.configure(
                    text=f"总进尺:{total_footage}m -> 费用:¥{raw_cost:,.0f}",
                    foreground=UI_COLORS["primary"],
                )
            else:
                self.current_cost = 0
                self.lbl_cost.configure(text=f"费用: ¥{raw_cost:,.0f} (未计入)", foreground=UI_COLORS["muted"])
        except:
            self.current_cost = 0
            self.lbl_cost.configure(text="地勘费用: 错误")
        if self.update_callback: self.update_callback()

    def get_cost(self):
        return getattr(self, "current_cost", 0)

class SurveyProcurementPanel(ttk.LabelFrame):
    def __init__(self, parent, update_callback):
        super().__init__(parent, text="5. 测绘采购成本 (New)", padding=5)
        self.update_callback = update_callback
        self.init_ui()
        
    def init_ui(self):
        self.v_enable = tk.BooleanVar(value=True)
        survey_header = ttk.Frame(self)
        ttk.Label(survey_header, text="5. 测绘采购成本").pack(side="left")
        ttk.Checkbutton(
            survey_header,
            text="计入总造价",
            variable=self.v_enable,
            command=self.calc,
        ).pack(side="left", padx=(8, 0))
        self.configure(text="")
        self.configure(labelwidget=survey_header)

        self.survey_scroll = SectionScrollArea(self, height=95, fit_width=True, with_hscroll=True)
        self.survey_scroll.pack(fill="both", expand=True)
        body = self.survey_scroll.body

        f_main = ttk.Frame(body)
        f_main.pack(fill="x", pady=5)
        self.e_area = LabelEntry(f_main, "测绘面积 (km²)", "30", width=10, callback=self.calc)
        self.e_area.pack(side="left", padx=2)
        
        ttk.Label(body, text="* 单价: ≤30km²:1600; ≤100km²:1300; >100km²:1100", style="Hint.TLabel").pack(anchor="w", padx=5)

        self.lbl_cost = ttk.Label(body, text="测绘费用: ¥0", style="Cost.TLabel")
        self.lbl_cost.pack(anchor="e", pady=5)
        # 初始化时不计算，避免覆盖默认值，或者在 set_area_by_capacity 后计算

    def set_area_by_capacity(self, capacity):
        """ 根据容量自动预估测绘面积 """
        if capacity <= 50: area = 30
        elif capacity <= 100: area = 60
        elif capacity <= 200: area = 130
        elif capacity <= 500: area = 250
        else: area = 400 # <= 1000 and > 1000
        
        self.e_area.set(area)
        self.calc()

    def calc(self, event=None):
        try:
            area = float(self.e_area.get())
        except:
            area = 0
        
        # 新的计费逻辑
        cost = 0
        if area > 0:
            if area <= 30:
                price = 1600
            elif area <= 100:
                price = 1300
            else:
                price = 1100
            cost = area * price

        if self.v_enable.get():
            self.current_cost = cost
            self.lbl_cost.configure(text=f"费用: ¥{cost:,.0f} (单价{price})", foreground=UI_COLORS["primary"])
        else:
            self.current_cost = 0
            self.lbl_cost.configure(text=f"费用: ¥{cost:,.0f} (未计入)", foreground=UI_COLORS["muted"])
        if self.update_callback: self.update_callback()

    def get_cost(self):
        return getattr(self, "current_cost", 0)


class OtherCostPanel(ttk.LabelFrame):
    def __init__(self, parent, update_callback):
        super().__init__(parent, text="6. 其他费用成本", padding=5)
        self.update_callback = update_callback
        self.capacity = 100.0
        self.init_ui()

    def init_ui(self):
        self.v_enable = tk.BooleanVar(value=True)
        header = ttk.Frame(self)
        ttk.Label(header, text="6. 其他费用成本").pack(side="left")
        ttk.Checkbutton(
            header,
            text="计入总造价",
            variable=self.v_enable,
            command=self.calc,
        ).pack(side="left", padx=(8, 0))
        self.configure(text="")
        self.configure(labelwidget=header)

        self.other_scroll = SectionScrollArea(self, height=130, fit_width=True, with_hscroll=True)
        self.other_scroll.pack(fill="both", expand=True)
        body = self.other_scroll.body

        f1 = ttk.Frame(body)
        f1.pack(fill="x", pady=(0, 2))
        self.v_consult = tk.BooleanVar(value=True)
        self.v_design = tk.BooleanVar(value=False)
        ttk.Checkbutton(f1, text="咨询（可研）阶段", variable=self.v_consult, command=self.calc).pack(side="left")
        ttk.Checkbutton(f1, text="设计阶段", variable=self.v_design, command=self.calc).pack(side="left", padx=(10, 0))

        f2 = ttk.Frame(body)
        f2.pack(fill="x", pady=(0, 2))
        self.v_design_stationed = tk.BooleanVar(value=False)
        self.v_design_international_mail = tk.BooleanVar(value=False)
        ttk.Checkbutton(f2, text="设计阶段含驻场", variable=self.v_design_stationed, command=self.calc).pack(side="left")
        ttk.Checkbutton(f2, text="图纸国际邮寄", variable=self.v_design_international_mail, command=self.calc).pack(side="left", padx=(10, 0))

        self.lbl_cost = ttk.Label(body, text="其他费用: ¥0", style="Cost.TLabel")
        self.lbl_cost.pack(anchor="e", pady=4)
        self.calc()

    def set_capacity(self, capacity):
        try:
            self.capacity = float(capacity)
        except:
            self.capacity = 0.0
        self.calc()

    def _get_consult_tier(self):
        cap = self.capacity
        if cap <= 100:
            return "le_100"
        if cap <= 200:
            return "101_200"
        if cap <= 500:
            return "201_500"
        return "gt_500"

    def _get_consult_detail_wan(self):
        tier = self._get_consult_tier()
        items = OTHER_COST_STD["consult_items_wan"]
        return {
            "交通": float(items["traffic"][tier]),
            "住宿": float(items["lodging"][tier]),
            "租车": float(items["rental"][tier]),
            "差补": float(items["subsidy"][tier]),
            "打印": float(items["printing"][tier]),
        }

    def _get_consult_cost_wan(self):
        return sum(self._get_consult_detail_wan().values())

    def _get_design_cost_wan(self):
        return sum(self._get_design_detail_wan().values())

    def _get_design_detail_wan(self):
        d = OTHER_COST_STD["design_items_wan"]
        cap = self.capacity
        detail = {
            "交通": float(d["traffic"]),
            "住宿": float(d["lodging_le_200"] if cap <= 200 else d["lodging_gt_200"]),
            "差补": float(d["subsidy_le_200"] if cap <= 200 else d["subsidy_gt_200"]),
            "打印": float(d["printing_international"] if self.v_design_international_mail.get() else d["printing_domestic"]),
        }
        if self.v_design_stationed.get():
            detail["驻场"] = float(d["stationed_le_200"] if cap <= 200 else d["stationed_gt_200"])
        return detail

    def calc(self, event=None):
        consult_detail = self._get_consult_detail_wan() if self.v_consult.get() else {}
        design_detail = self._get_design_detail_wan() if self.v_design.get() else {}
        consult_wan = self._get_consult_cost_wan() if self.v_consult.get() else 0.0
        design_wan = self._get_design_cost_wan() if self.v_design.get() else 0.0
        total_wan = consult_wan + design_wan
        raw_cost = total_wan * 10000.0

        if self.v_enable.get():
            self.current_cost = raw_cost
            self.lbl_cost.configure(text=f"其他费用: ¥{raw_cost:,.0f}", foreground=UI_COLORS["primary"])
        else:
            self.current_cost = 0.0
            self.lbl_cost.configure(text=f"其他费用: ¥{raw_cost:,.0f} (未计入)", foreground=UI_COLORS["muted"])

        if self.update_callback:
            self.update_callback()

    def get_cost(self):
        return getattr(self, "current_cost", 0.0)


class FinancialAnalysisPanel(ttk.LabelFrame):
    def __init__(self, parent, update_callback):
        super().__init__(parent, text="7. 利润分析", padding=5)
        self.update_callback = update_callback
        self.system_labor_cost = 0.0
        self.system_geo_cost = 0.0
        self.system_survey_cost = 0.0
        self.system_other_cost = 0.0
        self.last_metrics = {}
        self._init_ui()
        self.update_metrics(0.0, 0.0, 0.0, 0.0)

    def _init_ui(self):
        self.finance_scroll = SectionScrollArea(self, height=120, fit_width=True, with_hscroll=True)
        self.finance_scroll.pack(fill="both", expand=True)
        body = self.finance_scroll.body

        f1 = ttk.Frame(body)
        f1.pack(fill="x", pady=(0, 2))
        self.e_contract_tax = LabelEntry(f1, "合同额(含税,元)", "0", width=12, readonly=False, callback=self._on_manual_change)
        self.e_contract_tax.pack(side="left", padx=2)
        self.e_total_cost = LabelEntry(f1, "总成本(元)", "0", width=14, readonly=True)
        self.e_total_cost.pack(side="left", padx=2)

        self.lbl_sys = ttk.Label(body, text="人工: ¥0.00 | 外购(地勘+测绘): ¥0.00 | 差旅: ¥0.00", style="Hint.TLabel")
        self.lbl_sys.pack(anchor="w", pady=(2, 1))
        self.lbl_result = ttk.Label(body, text="不含税合同额: ¥0.00 | 利润率: --", style="Cost.TLabel")
        self.lbl_result.pack(anchor="w", pady=(1, 4))
        self.e_contract_tax.var.trace_add("write", lambda *_: self._recalc())

    def _parse_float(self, val, default=0.0):
        try:
            return float(str(val).replace(",", "").strip())
        except:
            return float(default)

    def _fmt_money(self, val):
        return f"¥{float(val):,.2f}"

    def _fmt_pct(self, val):
        if val is None:
            return "--"
        return f"{float(val) * 100:.2f}%"

    def _on_manual_change(self, event=None):
        self._recalc()
        if self.update_callback:
            self.update_callback()

    def update_metrics(self, labor_cost, geo_cost, survey_cost, other_cost=0.0):
        self.system_labor_cost = self._parse_float(labor_cost, 0.0)
        self.system_geo_cost = self._parse_float(geo_cost, 0.0)
        self.system_survey_cost = self._parse_float(survey_cost, 0.0)
        self.system_other_cost = self._parse_float(other_cost, 0.0)
        self._recalc()
        return self.get_metrics()

    def _recalc(self):
        contract_taxed = self._parse_float(self.e_contract_tax.get(), 0.0)
        labor_used = self.system_labor_cost
        external_system = self.system_geo_cost + self.system_survey_cost
        travel = self.system_other_cost
        base_cost = labor_used + external_system + travel
        emergency_reserve = base_cost * EMERGENCY_RESERVE_RATE
        total_cost = base_cost + emergency_reserve

        contract_no_tax = contract_taxed / 1.06 if contract_taxed > 0 else 0.0
        profit_rate = None
        if contract_no_tax > 0:
            profit_rate = (contract_no_tax - total_cost) / contract_no_tax

        self.e_total_cost.set(round(total_cost, 2))
        self.lbl_sys.configure(
            text=(
                f"人工: {self._fmt_money(labor_used)} | "
                f"外购(地勘+测绘): {self._fmt_money(external_system)} | "
                f"差旅: {self._fmt_money(travel)} | "
                f"应急储备: {self._fmt_money(emergency_reserve)}"
            ),
            foreground=UI_COLORS["primary"],
        )
        self.lbl_result.configure(
            text=(
                f"不含税合同额: {self._fmt_money(contract_no_tax)} | "
                f"利润率: {self._fmt_pct(profit_rate)}"
            ),
            foreground=UI_COLORS["primary"],
        )

        self.last_metrics = {
            "contract_taxed": contract_taxed,
            "contract_no_tax": contract_no_tax,
            "labor_used": labor_used,
            "external_system": external_system,
            "travel": travel,
            "base_cost": base_cost,
            "emergency_reserve": emergency_reserve,
            "total_cost": total_cost,
            "profit_rate": profit_rate,
        }

    def get_metrics(self):
        return dict(self.last_metrics)

    def set_contract_values(self, contract_taxed=None, **kwargs):
        if contract_taxed is not None:
            self.e_contract_tax.set(contract_taxed)
        self._recalc()

class RatioManagerWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("专业比例配置")
        self.geometry("400x300")
        self.configure(bg=UI_COLORS["bg"])
        f_info = ttk.LabelFrame(self, text="导入说明")
        f_info.pack(fill="x", padx=10, pady=10)
        ttk.Label(f_info, text="Excel需包含列: 专业, 主设人, 设计人, 校核人, 审核人").pack(padx=10, pady=5)
        ttk.Button(self, text="📂 导入专业比例配置", command=self.import_ratio_excel).pack(pady=20)
        self.lbl_status = ttk.Label(self, text=f"当前已配置专业数: {len(CURRENT_RATIO_DB)-1}")
        self.lbl_status.pack()
    def import_ratio_excel(self):
        if HAS_OPENPYXL:
            file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xlsm"), ("CSV Files", "*.csv")])
            if not file_path: return
            try:
                if file_path.lower().endswith(".csv"):
                    with open(file_path, "r", encoding="utf-8-sig") as f:
                        reader = csv.DictReader(f)
                        req_cols = ["专业", "主设人", "设计人", "校核人", "审核人"]
                        if not all(col in reader.fieldnames for col in req_cols):
                            return messagebox.showerror("错误", f"缺少列: {req_cols}")
                        rows = list(reader)
                else:
                    wb = load_workbook(file_path, data_only=True)
                    ws = wb.active
                    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
                    req_cols = ["专业", "主设人", "设计人", "校核人", "审核人"]
                    if not all(col in headers for col in req_cols):
                        return messagebox.showerror("错误", f"缺少列: {req_cols}")
                    rows = []
                    for r in ws.iter_rows(min_row=2, values_only=True):
                        row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
                        rows.append(row)

                count = 0
                for row in rows:
                    dept = str(row.get("专业", "")).strip()
                    if not dept: continue
                    try:
                        CURRENT_RATIO_DB[dept] = {
                            "主设人": float(row.get("主设人", 0) or 0),
                            "设计人": float(row.get("设计人", 0) or 0),
                            "校核人": float(row.get("校核人", 0) or 0),
                            "审核人": float(row.get("审核人", 0) or 0),
                        }
                        count += 1
                    except: continue
                save_json(RATIO_FILE, CURRENT_RATIO_DB); self.lbl_status.configure(text=f"当前已配置专业数: {len(CURRENT_RATIO_DB)-1}")
                messagebox.showinfo("成功", f"更新 {count} 个专业比例")
            except Exception as e: messagebox.showerror("失败", str(e))
        elif HAS_PANDAS:
            file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
            if not file_path: return
            try:
                with pd.ExcelFile(file_path) as xls: df = pd.read_excel(xls)
                df.columns = [str(c).strip() for c in df.columns]
                req_cols = ["专业", "主设人", "设计人", "校核人", "审核人"]
                if not all(col in df.columns for col in req_cols): return messagebox.showerror("错误", f"缺少列: {req_cols}")
                count = 0
                for _, row in df.iterrows():
                    dept = str(row["专业"]).strip()
                    try:
                        CURRENT_RATIO_DB[dept] = {"主设人": float(row["主设人"]), "设计人": float(row["设计人"]), "校核人": float(row["校核人"]), "审核人": float(row["审核人"])}
                        count += 1
                    except: continue
                save_json(RATIO_FILE, CURRENT_RATIO_DB); self.lbl_status.configure(text=f"当前已配置专业数: {len(CURRENT_RATIO_DB)-1}")
                messagebox.showinfo("成功", f"更新 {count} 个专业比例")
            except Exception as e: messagebox.showerror("失败", str(e))
        else:
            file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
            if not file_path: return
            try:
                with open(file_path, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    req_cols = ["专业", "主设人", "设计人", "校核人", "审核人"]
                    if not all(col in reader.fieldnames for col in req_cols):
                        return messagebox.showerror("错误", f"缺少列: {req_cols}")
                    count = 0
                    for row in reader:
                        dept = str(row.get("专业", "")).strip()
                        if not dept: continue
                        try:
                            CURRENT_RATIO_DB[dept] = {
                                "主设人": float(row.get("主设人", 0) or 0),
                                "设计人": float(row.get("设计人", 0) or 0),
                                "校核人": float(row.get("校核人", 0) or 0),
                                "审核人": float(row.get("审核人", 0) or 0),
                            }
                            count += 1
                        except: continue
                save_json(RATIO_FILE, CURRENT_RATIO_DB); self.lbl_status.configure(text=f"当前已配置专业数: {len(CURRENT_RATIO_DB)-1}")
                messagebox.showinfo("成功", f"更新 {count} 个专业比例")
            except Exception as e:
                messagebox.showerror("失败", str(e))

class SmartPasteDialog(tk.Toplevel):
    def __init__(self, parent, callback):
        super().__init__(parent)
        self.title("智能识别窗口")
        self.geometry("600x400")
        self.configure(bg=UI_COLORS["bg"])
        self.callback = callback
        ttk.Label(self, text="请将微信/邮件/Excel中的分工内容粘贴到下方，点击识别：").pack(padx=10, pady=5, anchor="w")
        ttk.Label(self, text="支持格式：'可行性研究报告：电气张三40%，李四30%' (自动分配角色和比例)").pack(padx=10, pady=0, anchor="w")
        self.txt = tk.Text(self, height=15)
        self.txt.pack(fill="both", expand=True, padx=10, pady=5)
        f_btn = ttk.Frame(self)
        f_btn.pack(fill="x", padx=10, pady=10)
        ttk.Button(f_btn, text="开始识别", command=self.on_ok).pack(side="right")
    def on_ok(self):
        text = self.txt.get("1.0", tk.END).strip()
        if text: self.callback(text)
        self.destroy()

class HelpDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("帮助说明")
        self.geometry("520x360")
        self.configure(bg=UI_COLORS["bg"])
        ttk.Label(self, text="帮助说明", font=("Microsoft YaHei UI", 12, "bold")).pack(padx=12, pady=(12, 6), anchor="w")
        txt = tk.Text(self, height=12, wrap="word")
        txt.pack(fill="both", expand=True, padx=12, pady=6)
        txt.insert("1.0", HELP_TEXT)
        txt.configure(state="disabled")


class RoleColumn:
    def __init__(self, parent, dept_name, role_name, pct, hours, update_cb, wage_override=None):
        self.dept_name = dept_name
        self.role_name = role_name
        self.hours = hours
        self.update_cb = update_cb

        if wage_override is None:
            self.wage = get_fixed_wage(dept_name, role_name)
        else:
            self.wage = float(wage_override)
        
        self.frame = ttk.Frame(parent, relief="flat", borderwidth=1)
        self.frame.pack(side="left", padx=5, pady=2, fill="y")
        
        ttk.Label(self.frame, text=role_name, font=UI_FONT_BOLD).pack(anchor="w")
        
        f_ratio = ttk.Frame(self.frame)
        f_ratio.pack(anchor="w", pady=1)
        self.ratio_var = tk.DoubleVar(value=pct)
        e_ratio = ttk.Entry(f_ratio, textvariable=self.ratio_var, width=3)
        e_ratio.pack(side="left")
        ttk.Label(f_ratio, text="%").pack(side="left")
        e_ratio.bind("<Return>", lambda e: self.trigger_recalc())
        
        # 移除姓名输入框，隐藏日薪显示 (Hide individual wage)
        
        days = hours / 8.0
        self.lbl_hours = ttk.Label(self.frame, text=f"{days:.1f}天 / {hours:.1f}h", style="Hint.TLabel")
        self.lbl_hours.pack(anchor="w")

    def trigger_recalc(self): self.update_cb(recalc_ratios=True)
    def is_valid(self): return True 
    def get_cost(self): return (self.hours / 8.0) * self.wage
    def get_ratio(self):
        try: return self.ratio_var.get()
        except: return 0.0
    def set_profile(self, pct, hours, wage=None):
        try:
            self.ratio_var.set(float(pct))
        except:
            self.ratio_var.set(0.0)
        if wage is not None:
            self.wage = float(wage)
        self.set_hours(hours)
    def set_hours(self, h): 
        self.hours = h
        days = h / 8.0
        self.lbl_hours.configure(text=f"{days:.1f}天 / {h:.1f}h")

class DisciplineRow:
    def __init__(self, parent, dept_name, total_days, init_ratios, update_project_total_cb, wage_overrides=None):
        self.dept_name = dept_name
        self.total_days = total_days
        self.update_project_total_cb = update_project_total_cb
        self.wage_overrides = wage_overrides or {}
        
        self.frame = ttk.LabelFrame(parent, text=f"📂 {dept_name} (计算中...)")
        self.inner_frame = ttk.Frame(self.frame)
        self.inner_frame.pack(fill="x", padx=2, pady=2)
        self.role_cols = []
        self.role_map = {}
        
        for role in SYSTEM_ROLE_ORDER:
            col = RoleColumn(
                self.inner_frame,
                dept_name,
                role,
                0.0,
                0.0,
                self.update_me,
                wage_override=self.wage_overrides.get(role, get_fixed_wage(dept_name, role)),
            )
            self.role_cols.append(col)
            self.role_map[role] = col
        self.apply_profile(total_days, init_ratios, wage_overrides=self.wage_overrides)

    def _resolve_ratios(self, init_ratios):
        my_ratios = init_ratios.copy()
        for k, v in CURRENT_RATIO_DB.items():
            if k in self.dept_name:
                my_ratios = v
                break
        return my_ratios

    def apply_profile(self, total_days, init_ratios, wage_overrides=None):
        self.total_days = total_days
        if wage_overrides is not None:
            self.wage_overrides = wage_overrides or {}

        my_ratios = self._resolve_ratios(init_ratios)
        for role in SYSTEM_ROLE_ORDER:
            col = self.role_map[role]
            pct = float(my_ratios.get(role, 0) or 0)
            role_hours = total_days * (pct / 100.0) * 8
            override_wage = self.wage_overrides.get(role)
            wage = get_fixed_wage(self.dept_name, role) if override_wage is None else float(override_wage)
            col.set_profile(pct, role_hours, wage=wage)

        self._set_title(sum(col.get_cost() for col in self.role_cols))
            
    def is_ready(self): return True
    def get_total_hours(self): return sum(col.hours for col in self.role_cols)
    def _set_title(self, total_cost):
        title = (
            f"📂 {self.dept_name} "
            f"({self.total_days:.1f}工日 / {self.total_days*8:.1f}工时 | ¥{total_cost:,.0f})"
        )
        self.frame.configure(text=title)
    def update_me(self, recalc_ratios=False):
        if recalc_ratios:
            for col in self.role_cols:
                col.set_hours(self.total_days * (col.get_ratio()/100.0) * 8)
        current_total = sum(col.get_cost() for col in self.role_cols)
        self._set_title(current_total)
        self.update_project_total_cb()
    def get_total_cost(self): return sum(col.get_cost() for col in self.role_cols)

# ==========================================
# 5. 主程序 App (集成导出功能)
# ==========================================
class App:
    def __init__(self, root):
        self.root = root
        apply_theme(self.root)
        self.root.title("瑞科设计成本核算系统 V95.0 (Survey Area Auto-Link)")
        if platform.system() == "Windows": self.root.state('zoomed')
        else: self.root.geometry("1400x900")

        self.dept_rows = {}
        self.task_breakdown_rows = []
        self.geo_panel = None
        self.survey_panel = None
        self.other_panel = None
        self.finance_panel = None
        self.last_finance_metrics = {}
        self.project_staff_assignments = {}
        self.project_staff_role_wages = {}
        self.project_staff_match_details = []
        self.project_staff_valid_count = 0
        self.project_staff_skip_count = 0
        self.project_staff_import_file = ""
        self.project_staff_import_time = ""
        self.project_staff_last_block_reason = ""
        self.v_labor_pricing_mode = tk.StringVar(value="avg")
        self._calc_total_job = None

        self.lbl_total = ttk.Label(root, text="总成本: (等待计算)", style="Summary.TLabel")
        self.lbl_total.pack(side="bottom", fill="x", pady=8, padx=10)
        
        paned = ttk.PanedWindow(root, orient=tk.HORIZONTAL)
        paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        f_left = ttk.Frame(paned, width=380)
        paned.add(f_left, weight=1)
        # 左侧增加纵向可拖拽分隔，支持上下拉伸/压缩
        self.left_vertical_paned = ttk.PanedWindow(f_left, orient=tk.VERTICAL)
        self.left_vertical_paned.pack(fill="both", expand=True)
        f_left_top = ttk.Frame(self.left_vertical_paned)
        f_left_bottom = ttk.Frame(self.left_vertical_paned)
        self.left_vertical_paned.add(f_left_top, weight=4)
        self.left_vertical_paned.add(f_left_bottom, weight=2)

        # --- Logo Section ---
        self.logo_frame = ttk.Frame(f_left_top)
        self.logo_frame.pack(fill="x", padx=5, pady=5)

        icon_path = resource_path("LOGO.ico")
        if os.path.exists(icon_path):
            try: self.root.iconbitmap(icon_path)
            except: pass

        lbl_title = ttk.Label(self.logo_frame, text="瑞科设计成本核算系统", font=UI_FONT_TITLE, foreground=UI_COLORS["text"])
        lbl_title.pack(pady=10)

        
        # 1. 基础参数
        lf1 = ttk.LabelFrame(f_left_top, text="1. 基础参数")
        lf1.pack(fill="x", padx=5, pady=5)
        self.v_type = tk.StringVar(value="风电")
        ttk.Radiobutton(lf1, text="风电", variable=self.v_type, value="风电", command=self.on_project_change).pack(side="left")
        ttk.Radiobutton(lf1, text="光伏", variable=self.v_type, value="光伏", command=self.on_project_change).pack(side="left")
        ttk.Radiobutton(lf1, text="储能", variable=self.v_type, value="储能", command=self.on_project_change).pack(side="left")
        
        self.e_cap = ttk.Entry(lf1, width=5)
        self.e_cap.pack(side="left", padx=5)
        self.e_cap.insert(0,"100")
        # 绑定回车和焦点离开事件，自动更新测绘面积
        self.e_cap.bind("<Return>", self.on_cap_change)
        self.e_cap.bind("<FocusOut>", self.on_cap_change)

        ttk.Label(lf1, text="MW").pack(side="left")
        
        self.c_intl = ttk.Combobox(lf1, width=28, state="readonly", values=[
            "1.0 (非国际项目)", 
            "1.2 (国标)",
            "1.5 (国标+国外业主)",
            "2.0 (国际标准)"
        ])
        self.c_intl.current(0); self.c_intl.pack(side="left", padx=5)

        lf1_mode = ttk.Frame(f_left_top)
        lf1_mode.pack(fill="x", padx=5, pady=(0, 5))
        ttk.Label(lf1_mode, text="人工计价方式：").pack(side="left")
        ttk.Radiobutton(
            lf1_mode,
            text="专业未定（平均日薪）",
            variable=self.v_labor_pricing_mode,
            value="avg",
            command=self.on_labor_pricing_mode_change,
        ).pack(side="left")
        ttk.Radiobutton(
            lf1_mode,
            text="导入项目人员名单",
            variable=self.v_labor_pricing_mode,
            value="staff",
            command=self.on_labor_pricing_mode_change,
        ).pack(side="left", padx=(10, 0))

        # 2. 默认比例 (5, 71, 16, 8)
        lf2 = ttk.LabelFrame(f_left_top, text="2. 全局比例 (%)")
        lf2.pack(fill="x", padx=5, pady=5)
        self.vars_ratio = {}
        defaults = [("主设",5), ("设计",71), ("校核",16), ("审核",8)]
        for i, (lbl, val) in enumerate(defaults):
            ttk.Label(lf2, text=lbl).grid(row=0, column=i)
            v = tk.DoubleVar(value=val)
            ttk.Entry(lf2, textvariable=v, width=3).grid(row=1, column=i)
            self.vars_ratio[lbl] = v

        # 3. 任务选择
        lf3 = ttk.LabelFrame(f_left_top, text="3. 任务选择 (多选)")
        lf3.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.v_ext_micro = tk.BooleanVar()
        self.v_ext_turbine = tk.BooleanVar()
        self.v_ext_proposal = tk.BooleanVar()
        self.v_ext_app = tk.BooleanVar()
        self.v_ext_bid = tk.BooleanVar()
        self.v_grid_review = tk.BooleanVar()
        self.v_has_outgoing = tk.BooleanVar()
        self.v_has_line_survey = tk.BooleanVar()
        self.v_storage_bundle = tk.BooleanVar()
        self.v_prelim_micro_bonus = tk.BooleanVar()
        self.v_line_survey_terrain = tk.StringVar(value="非山地")
        self.v_booster_kv = tk.StringVar(value="常规")
        lf_extras = ttk.LabelFrame(lf3, text="其他修订项")
        lf_extras.pack(fill="x", padx=5, pady=2)
        f_extra_tools = ttk.Frame(lf_extras)
        f_extra_tools.pack(fill="x", padx=4, pady=4)
        ttk.Button(f_extra_tools, text="点开选择其他修订项", command=self.open_extras_selector).pack(side="left")
        self.lbl_extras_selected = ttk.Label(f_extra_tools, text="已选修订项: 0 项", style="Hint.TLabel")
        self.lbl_extras_selected.pack(side="left", padx=(10, 0))

        f_tools = ttk.Frame(lf3)
        f_tools.pack(fill="x")
        ttk.Button(f_tools, text="点开选择任务", command=self.open_task_selector).pack(side="left")
        ttk.Button(f_tools, text="全选", command=self.select_all_tasks).pack(side="left", padx=(6, 0))
        ttk.Button(f_tools, text="清空", command=self.clear_all_tasks).pack(side="left", padx=(6, 0))
        self.lbl_task_selected = ttk.Label(f_tools, text="已选任务: 0 项", style="Hint.TLabel")
        self.lbl_task_selected.pack(side="right")

        # 任务勾选状态容器（不在主界面展开显示，改为弹窗确认选择）
        self.chk_frame = ScrollableCheckBoxFrame(lf3, height=1)

        # 左下采购区：使用滚动容器，确保所有模块可见
        f_procurement_scroll = SectionScrollArea(f_left_bottom, height=320, fit_width=True, with_hscroll=False)
        f_procurement_scroll.pack(fill="both", expand=True, padx=5, pady=5)
        f_procurement = f_procurement_scroll.body

        self.geo_panel = GeotechProcurementPanel(f_procurement, self.request_calc_total)
        self.geo_panel.pack(fill="x", pady=2)
        
        self.survey_panel = SurveyProcurementPanel(f_procurement, self.request_calc_total)
        self.survey_panel.pack(fill="x", pady=2)

        self.other_panel = OtherCostPanel(f_procurement, self.request_calc_total)
        self.other_panel.pack(fill="x", pady=2)

        self.finance_panel = FinancialAnalysisPanel(f_procurement, self.request_calc_total)
        self.finance_panel.pack(fill="x", pady=2)

        f_right = ttk.Frame(paned)
        paned.add(f_right, weight=3)
        
        f_r_tools = ttk.Frame(f_right)
        f_r_tools.pack(fill="x", pady=5)
        
        btn_gen = ttk.Button(f_r_tools, text="👉 生成人工预算表", command=self.generate_report, style="Primary.TButton")
        btn_gen.pack(side="left", padx=5)
        ttk.Separator(f_r_tools, orient="vertical").pack(side="left", fill="y", padx=5)
        
        # 导出按钮 (New!)
        btn_export = ttk.Button(f_r_tools, text="📥 导出精美报表", command=self.export_excel, style="Primary.TButton")
        btn_export.pack(side="left", padx=5)

        ttk.Button(f_r_tools, text="👥 导入项目人员名单", command=self.import_project_staff).pack(side="left", padx=5)
        ttk.Button(f_r_tools, text="📄 项目人员模板", command=self.download_project_staff_template).pack(side="left", padx=5)
        ttk.Button(f_r_tools, text="📋 智能识别", command=self.open_smart_paste).pack(side="left", padx=5)
        ttk.Button(f_r_tools, text="⚙️ 专业比例管理", command=self.open_ratio_mgr).pack(side="left", padx=5)
        ttk.Button(f_r_tools, text="📥 导出模板", command=self.download_template).pack(side="left", padx=5)
        ttk.Button(f_r_tools, text="❓ 帮助说明", command=self.open_help).pack(side="right", padx=5)
        self.lbl_staff_status = ttk.Label(f_right, text="人员计价状态：未导入项目人员名单", style="Hint.TLabel")
        self.lbl_staff_status.pack(fill="x", padx=6, pady=(0, 4), anchor="w")

        self.canvas = tk.Canvas(f_right, bg=UI_COLORS["bg"], highlightthickness=0) 
        self.v_scroll = ttk.Scrollbar(f_right, orient="vertical", command=self.canvas.yview)
        self.scroll_frame = ttk.Frame(self.canvas)
        
        self.scroll_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window_content = self.canvas.create_window((0,0), window=self.scroll_frame, anchor="nw")
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_window_content, width=e.width))

        self.canvas.configure(yscrollcommand=self.v_scroll.set)
        self.v_scroll.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.refresh_list()
        self.on_project_change()
        self.generate_report(silent_if_missing_staff=True)
        # 初始化时触发一次容量更新，确保测绘面积正确
        self.on_cap_change()
        self.refresh_staff_status()

    def on_cap_change(self, event=None):
        """ 容量改变时，自动更新测绘面积 """
        try:
            cap = float(self.e_cap.get())
            if self.survey_panel:
                self.survey_panel.set_area_by_capacity(cap)
            if self.other_panel:
                self.other_panel.set_capacity(cap)
        except: pass

    def request_calc_total(self):
        # 输入过程中统一做轻量防抖，减少每次按键都触发整页刷新导致的卡顿。
        if not hasattr(self, "root"):
            return
        if self._calc_total_job is not None:
            try:
                self.root.after_cancel(self._calc_total_job)
            except Exception:
                pass
        self._calc_total_job = self.root.after(120, self._run_calc_total_from_timer)

    def _run_calc_total_from_timer(self):
        self._calc_total_job = None
        self.calc_total(show_warning_if_zero=False)

    def on_project_change(self):
        self.refresh_list()
        p_type = self.v_type.get()
        if p_type != "储能":
            self.v_storage_bundle.set(False)
        self._update_selected_extras_summary()
        if self.geo_panel:
            self.geo_panel.set_project_type(p_type)

    def get_runtime_profile(self):
        p_type = self.v_type.get()
        if p_type == "风电":
            return DATA_WIND_PARAMS, DB_WIND_PARAMS, MATRIX_WIND
        if p_type == "光伏":
            return DATA_PV_PARAMS, DB_PV_PARAMS, MATRIX_PV
        # 储能沿用光伏侧基础库，但在任务列表上做储能过滤
        return DATA_PV_PARAMS, DB_PV_PARAMS, MATRIX_PV

    def get_visible_task_keys(self):
        data, _, _ = self.get_runtime_profile()
        keys = [item["key"] for item in data]
        p_type = self.v_type.get()
        if p_type == "风电":
            keys = [k for k in keys if "（光伏项目，含升压站）" not in k and "（储能电站）" not in k]
        elif p_type == "光伏":
            keys = [k for k in keys if "（风电项目，含升压站）" not in k and "（储能电站）" not in k]
        elif p_type == "储能":
            keys = [k for k in keys if "（风电项目，含升压站）" not in k and "（光伏项目，含升压站）" not in k and "（升压电站）" not in k]
        # 保持顺序去重
        return list(dict.fromkeys(keys))

    def refresh_list(self):
        prev_selected = set(self.chk_frame.get_checked_items()) if hasattr(self, "chk_frame") else set()
        visible_items = self.get_visible_task_keys()
        self.chk_frame.populate(visible_items)
        for task in visible_items:
            if task in prev_selected:
                self.chk_frame.vars[task].set(True)
        if visible_items and not self.chk_frame.get_checked_items():
            self.chk_frame.vars[visible_items[0]].set(True)
        self._update_selected_task_summary()

    def _update_selected_task_summary(self):
        if not hasattr(self, "lbl_task_selected"):
            return
        selected = self.chk_frame.get_checked_items() if hasattr(self, "chk_frame") else []
        if not selected:
            self.lbl_task_selected.configure(text="已选任务: 0 项")
            return
        preview = "、".join(selected[:2])
        suffix = "" if len(selected) <= 2 else "..."
        self.lbl_task_selected.configure(text=f"已选任务: {len(selected)} 项（{preview}{suffix}）")

    def _get_extras_state(self):
        return {
            "ext_micro": self.v_ext_micro.get(),
            "ext_turbine": self.v_ext_turbine.get(),
            "ext_proposal": self.v_ext_proposal.get(),
            "ext_app": self.v_ext_app.get(),
            "ext_bid": self.v_ext_bid.get(),
            "grid_review": self.v_grid_review.get(),
            "has_outgoing": self.v_has_outgoing.get(),
            "has_line_survey": self.v_has_line_survey.get(),
            "line_survey_terrain": self.v_line_survey_terrain.get().strip() or "非山地",
            "booster_kv": self.v_booster_kv.get().strip() or "常规",
            "storage_bundle": self.v_storage_bundle.get(),
            "prelim_micro_bonus": self.v_prelim_micro_bonus.get(),
        }

    def _apply_extras_state(self, state):
        if not isinstance(state, dict):
            return
        self.v_ext_micro.set(bool(state.get("ext_micro", False)))
        self.v_ext_turbine.set(bool(state.get("ext_turbine", False)))
        self.v_ext_proposal.set(bool(state.get("ext_proposal", False)))
        self.v_ext_app.set(bool(state.get("ext_app", False)))
        self.v_ext_bid.set(bool(state.get("ext_bid", False)))
        self.v_grid_review.set(bool(state.get("grid_review", False)))
        self.v_has_outgoing.set(bool(state.get("has_outgoing", False)))
        self.v_has_line_survey.set(bool(state.get("has_line_survey", False)))
        self.v_line_survey_terrain.set(state.get("line_survey_terrain", "非山地") or "非山地")
        self.v_booster_kv.set(state.get("booster_kv", "常规") or "常规")
        storage_bundle = bool(state.get("storage_bundle", False))
        if self.v_type.get() != "储能":
            storage_bundle = False
        self.v_storage_bundle.set(storage_bundle)
        self.v_prelim_micro_bonus.set(bool(state.get("prelim_micro_bonus", False)))
        self._update_selected_extras_summary()

    def _update_selected_extras_summary(self):
        if not hasattr(self, "lbl_extras_selected"):
            return
        labels = []
        if self.v_ext_micro.get():
            labels.append("微观选址")
        if self.v_ext_turbine.get():
            labels.append("机型比选")
        if self.v_ext_proposal.get():
            labels.append("建议书")
        if self.v_ext_app.get():
            labels.append("项目申请")
        if self.v_ext_bid.get():
            labels.append("竞配报告")
        if self.v_grid_review.get():
            labels.append("配合电网评审")
        if self.v_has_outgoing.get():
            labels.append("含外送")
        if self.v_has_line_survey.get():
            labels.append(f"含线路测绘({self.v_line_survey_terrain.get().strip() or '非山地'})")
        booster_kv = self.v_booster_kv.get().strip() or "常规"
        if booster_kv != "常规":
            labels.append(f"升压站{booster_kv}")
        if self.v_storage_bundle.get():
            labels.append("若为配储")
        if self.v_prelim_micro_bonus.get():
            labels.append("初设微选附加")

        if not labels:
            self.lbl_extras_selected.configure(text="已选修订项: 0 项")
            return
        preview = "、".join(labels[:2])
        suffix = "" if len(labels) <= 2 else "..."
        self.lbl_extras_selected.configure(text=f"已选修订项: {len(labels)} 项（{preview}{suffix}）")

    def select_all_tasks(self):
        self.chk_frame.select_all()
        self._update_selected_task_summary()

    def clear_all_tasks(self):
        self.chk_frame.deselect_all()
        self._update_selected_task_summary()

    def open_task_selector(self):
        visible_items = self.get_visible_task_keys()
        preselected = set(self.chk_frame.get_checked_items())

        def _apply(selected_items):
            selected_set = set(selected_items or [])
            self.chk_frame.populate(visible_items)
            for task, var in self.chk_frame.vars.items():
                var.set(task in selected_set)
            if visible_items and not self.chk_frame.get_checked_items():
                self.chk_frame.vars[visible_items[0]].set(True)
            self._update_selected_task_summary()

        dlg = TaskSelectionDialog(self.root, visible_items, preselected=preselected, on_confirm=_apply)
        self.root.wait_window(dlg)

    def open_extras_selector(self):
        state = self._get_extras_state()

        def _apply(new_state):
            self._apply_extras_state(new_state)

        dlg = ExtrasSelectionDialog(
            self.root,
            state=state,
            project_type=self.v_type.get(),
            on_confirm=_apply,
        )
        self.root.wait_window(dlg)
     
    def open_smart_paste(self): SmartPasteDialog(self.root, self.process_smart_text)

    def open_ratio_mgr(self): RatioManagerWindow(self.root)
    def open_help(self): HelpDialog(self.root)

    def on_labor_pricing_mode_change(self):
        self.refresh_staff_status()
        self.generate_report(silent_if_missing_staff=True)

    def _set_staff_status(self, text):
        if hasattr(self, "lbl_staff_status"):
            self.lbl_staff_status.configure(text=text)

    def refresh_staff_status(self):
        mode = self.v_labor_pricing_mode.get() if hasattr(self, "v_labor_pricing_mode") else "avg"
        if mode == "avg":
            self._set_staff_status("人员计价状态：平均日薪模式（使用专业主设/设计/校核/审核平均日薪）")
            return

        if not self.project_staff_assignments:
            self._set_staff_status("人员计价状态：项目人员名单模式（未导入名单，缺失专业将自动跳过）")
            return

        if self.project_staff_last_block_reason:
            self._set_staff_status(f"人员计价状态：项目人员名单模式（{self.project_staff_last_block_reason}）")
            return

        self._set_staff_status(
            f"人员计价状态：项目人员名单模式（有效{self.project_staff_valid_count}人，跳过{self.project_staff_skip_count}人）"
        )

    def _read_project_staff_rows(self, file_path):
        rows = []
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                hmap = {clean_excel_header(h): h for h in headers}
                missing = [x for x in PROJECT_STAFF_TEMPLATE_HEADERS if clean_excel_header(x) not in hmap]
                if missing:
                    raise ValueError(f"缺少列: {missing}")
                for row in reader:
                    rows.append({col: row.get(hmap[clean_excel_header(col)], "") for col in PROJECT_STAFF_TEMPLATE_HEADERS})
            return rows

        if HAS_OPENPYXL:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            hmap = {clean_excel_header(h): idx for idx, h in enumerate(headers)}
            missing = [x for x in PROJECT_STAFF_TEMPLATE_HEADERS if clean_excel_header(x) not in hmap]
            if missing:
                raise ValueError(f"缺少列: {missing}")
            for r in ws.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for col in PROJECT_STAFF_TEMPLATE_HEADERS:
                    idx = hmap[clean_excel_header(col)]
                    val = r[idx] if idx < len(r) else ""
                    row_data[col] = "" if val is None else str(val).strip()
                rows.append(row_data)
            return rows

        if HAS_PANDAS:
            df = pd.read_excel(file_path)
            df.columns = [str(c).strip() for c in df.columns]
            hmap = {clean_excel_header(h): h for h in df.columns}
            missing = [x for x in PROJECT_STAFF_TEMPLATE_HEADERS if clean_excel_header(x) not in hmap]
            if missing:
                raise ValueError(f"缺少列: {missing}")
            for _, r in df.iterrows():
                row_data = {}
                for col in PROJECT_STAFF_TEMPLATE_HEADERS:
                    raw = r.get(hmap[clean_excel_header(col)], "")
                    row_data[col] = "" if pd.isna(raw) else str(raw).strip()
                rows.append(row_data)
            return rows

        raise ValueError("当前环境无法读取该文件（缺少 openpyxl/pandas）")

    def _build_project_staff_wages(self, assignment_map):
        details = []
        role_wages = {}
        valid_count = 0
        skip_count = 0

        for dept in sorted(assignment_map.keys(), key=get_dept_sort_index):
            role_wages[dept] = {}
            for role in SYSTEM_ROLE_ORDER:
                names = list(assignment_map.get(dept, {}).get(role, []))
                role_valid_wages = []
                for raw_name in names:
                    name = normalize_person_name(raw_name)
                    if not name:
                        continue

                    person_list = FIXED_PERSON_LOOKUP.get(name, [])
                    if not person_list:
                        # 特殊识别：支持“张伟 线路”“李乐乐-内蒙”等带空格/连字符写法。
                        for key in person_name_keys(name):
                            person_list = FIXED_PERSON_LOOKUP_NORM.get(key, [])
                            if person_list:
                                break
                    if not person_list:
                        skip_count += 1
                        details.append(
                            {
                                "专业": dept,
                                "角色": role,
                                "填报姓名": name,
                                "匹配姓名": "",
                                "区域": "",
                                "岗职": "",
                                "匹配日薪": "",
                                "状态": "跳过",
                                "跳过原因": "人员主数据不存在",
                            }
                        )
                        continue
                    if len(person_list) > 1:
                        skip_count += 1
                        details.append(
                            {
                                "专业": dept,
                                "角色": role,
                                "填报姓名": name,
                                "匹配姓名": "",
                                "区域": "",
                                "岗职": "",
                                "匹配日薪": "",
                                "状态": "跳过",
                                "跳过原因": "人员主数据同名冲突",
                            }
                        )
                        continue

                    person = person_list[0]
                    person_region = normalize_region_name(person.get("region", ""))
                    person_grade = person.get("grade", "")

                    # 口径：识别人员“区域+专业+岗职”并据此匹配日薪。
                    person_disc = str(person.get("discipline", "")).strip()
                    if not person_disc:
                        skip_count += 1
                        details.append(
                            {
                                "专业": dept,
                                "角色": role,
                                "填报姓名": name,
                                "匹配姓名": person.get("name", ""),
                                "区域": person_region,
                                "岗职": person_grade,
                                "匹配日薪": "",
                                "状态": "跳过",
                                "跳过原因": "人员主数据缺少专业",
                            }
                        )
                        continue

                    wage = FIXED_WAGE_LOOKUP.get((person_region, person_disc, person_grade))
                    if wage is None:
                        # 兜底1：同专业同岗级跨区域均值
                        wage = FIXED_WAGE_GROUP_AVG.get((person_disc, person_grade))
                    if wage is None:
                        # 兜底2：同区域同岗级跨专业均值
                        wage = FIXED_WAGE_REGION_GRADE_AVG.get((person_region, person_grade))
                    if wage is None:
                        # 兜底3：同岗级全库均值
                        wage = FIXED_WAGE_GRADE_AVG.get(person_grade)
                    if wage is None:
                        skip_count += 1
                        details.append(
                            {
                                "专业": dept,
                                "角色": role,
                                "填报姓名": name,
                                "匹配姓名": person.get("name", ""),
                                "区域": person_region,
                                "岗职": person_grade,
                                "匹配日薪": "",
                                "状态": "跳过",
                                "跳过原因": "固定日薪库缺少匹配项",
                            }
                        )
                        continue

                    role_valid_wages.append(float(wage))
                    valid_count += 1
                    details.append(
                        {
                            "专业": dept,
                            "角色": role,
                            "填报姓名": name,
                            "匹配姓名": person.get("name", ""),
                            "区域": person_region,
                            "岗职": person_grade,
                            "匹配日薪": round(float(wage), 2),
                            "状态": "有效",
                            "跳过原因": "",
                        }
                    )

                if role_valid_wages:
                    role_wages[dept][role] = sum(role_valid_wages) / len(role_valid_wages)

        return role_wages, details, valid_count, skip_count

    def _validate_required_staff_wages(self, required_depts):
        missing = []
        for dept in sorted(required_depts, key=get_dept_sort_index):
            role_map = self.project_staff_role_wages.get(dept, {})
            for role in SYSTEM_ROLE_ORDER:
                if role_map.get(role) is None:
                    missing.append((dept, role))
        return missing

    def import_project_staff(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV Files", "*.xlsx;*.xlsm;*.csv"), ("Excel Files", "*.xlsx;*.xlsm"), ("CSV Files", "*.csv")]
        )
        if not file_path:
            return
        try:
            raw_rows = self._read_project_staff_rows(file_path)
            assignment_map = {}
            unknown_disciplines = set()

            for row in raw_rows:
                raw_disc = str(row.get("专业", "")).strip()
                if not raw_disc:
                    continue
                dept = normalize_system_discipline_name(raw_disc)
                if not dept:
                    unknown_disciplines.add(raw_disc)
                    continue
                role_map = assignment_map.setdefault(dept, {role: [] for role in SYSTEM_ROLE_ORDER})
                for role in SYSTEM_ROLE_ORDER:
                    role_map[role].extend(split_person_names(row.get(role, "")))

            if unknown_disciplines:
                raise ValueError(f"存在无法识别的专业: {sorted(unknown_disciplines)}")
            if not assignment_map:
                raise ValueError("导入文件中没有有效的专业人员数据")

            # 每个专业+角色去重（保持原顺序）
            for dept, role_map in assignment_map.items():
                for role in SYSTEM_ROLE_ORDER:
                    role_map[role] = dedupe_person_names(role_map.get(role, []))

            role_wages, details, valid_count, skip_count = self._build_project_staff_wages(assignment_map)

            self.project_staff_assignments = assignment_map
            self.project_staff_role_wages = role_wages
            self.project_staff_match_details = details
            self.project_staff_valid_count = valid_count
            self.project_staff_skip_count = skip_count
            self.project_staff_import_file = file_path
            self.project_staff_import_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.project_staff_last_block_reason = ""

            self.refresh_staff_status()
            messagebox.showinfo("成功", f"项目人员名单导入完成。\n有效匹配：{valid_count} 人\n跳过：{skip_count} 人")
            self.generate_report()
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def process_smart_text(self, text):
        visible_keys = self.get_visible_task_keys()
        visible_key_set = set(visible_keys)
        found_tasks = []
        for task_name in visible_keys:
            if task_name in text:
                found_tasks.append(task_name)
        for key, aliases in TASK_SYNONYMS.items():
            if not isinstance(aliases, list): aliases = [aliases]
            if key in text:
                for alias in aliases:
                    alias = canonical_task_key(alias)
                    if alias in visible_key_set and alias not in found_tasks:
                        found_tasks.append(alias)
        if found_tasks:
            self.chk_frame.deselect_all()
            for task_name in found_tasks:
                if task_name in self.chk_frame.vars: self.chk_frame.vars[task_name].set(True)
            self._update_selected_task_summary()
            self.generate_report()
            self.root.update()
        else:
            if not self.chk_frame.get_checked_items() and not self.dept_rows:
                 return messagebox.showwarning("提示", "未识别到任务")
        self.calc_total()
        self._update_selected_task_summary()
        messagebox.showinfo("完成", "识别处理完毕")

    def download_project_staff_template(self):
        template_rows = []
        for dept in DISCIPLINE_ORDER:
            template_rows.append(
                {
                    "专业": dept,
                    "主设人": "",
                    "设计人": "",
                    "校核人": "",
                    "审核人": "",
                }
            )

        if HAS_OPENPYXL:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="项目人员名单模板.xlsx",
            )
            if not file_path:
                return
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "项目人员名单模板"
                ws.append(PROJECT_STAFF_TEMPLATE_HEADERS)
                for row in template_rows:
                    ws.append([row["专业"], row["主设人"], row["设计人"], row["校核人"], row["审核人"]])
                ws.column_dimensions["A"].width = 16
                ws.column_dimensions["B"].width = 18
                ws.column_dimensions["C"].width = 18
                ws.column_dimensions["D"].width = 18
                ws.column_dimensions["E"].width = 18
                wb.save(file_path)
                messagebox.showinfo("成功", f"项目人员模板已保存：\n{file_path}")
            except Exception as e:
                messagebox.showerror("失败", str(e))
            return

        if HAS_PANDAS:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="项目人员名单模板.xlsx",
            )
            if not file_path:
                return
            try:
                pd.DataFrame(template_rows).to_excel(file_path, index=False)
                messagebox.showinfo("成功", f"项目人员模板已保存：\n{file_path}")
            except Exception as e:
                messagebox.showerror("失败", str(e))
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            initialfile="项目人员名单模板.csv",
        )
        if not file_path:
            return
        try:
            write_csv(file_path, PROJECT_STAFF_TEMPLATE_HEADERS, template_rows)
            messagebox.showinfo("成功", f"项目人员模板已保存：\n{file_path}")
        except Exception as e:
            messagebox.showerror("失败", str(e))

    def download_template(self):
        all_depts = set()
        for matrix in [MATRIX_WIND, MATRIX_PV]:
            for task_depts in matrix.values(): all_depts.update(task_depts.keys())
        sorted_depts = sorted(list(all_depts), key=get_dept_sort_index)

        if HAS_OPENPYXL:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if not file_path: return
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "专业比例模板"
                headers = ["专业", "主设人", "设计人", "校核人", "审核人"]
                ws.append(headers)
                for d in sorted_depts:
                    ws.append([d, "", "", "", ""])
                for col, width in zip(["A", "B", "C", "D", "E"], [18, 10, 10, 10, 10]):
                    ws.column_dimensions[col].width = width
                wb.save(file_path)
                messagebox.showinfo("成功", "模板已保存")
            except Exception as e:
                messagebox.showerror("失败", str(e))
        elif HAS_PANDAS:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if not file_path: return
            data = {"专业": sorted_depts, "主设人": [""]*len(sorted_depts), "设计人": [""]*len(sorted_depts), "校核人": [""]*len(sorted_depts), "审核人": [""]*len(sorted_depts)}
            try: pd.DataFrame(data).to_excel(file_path, index=False); messagebox.showinfo("成功", "模板已保存")
            except Exception as e: messagebox.showerror("失败", str(e))
        else:
            file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
            if not file_path: return
            headers = ["专业", "主设人", "设计人", "校核人", "审核人"]
            rows = [{"专业": d, "主设人": "", "设计人": "", "校核人": "", "审核人": ""} for d in sorted_depts]
            try:
                write_csv(file_path, headers, rows)
                messagebox.showinfo("成功", f"CSV 模板已保存：\n{file_path}")
            except Exception as e:
                messagebox.showerror("失败", str(e))

    def generate_report(self, silent_if_missing_staff=False):
        selected = self.chk_frame.get_checked_items()
        if not selected: 
            self.chk_frame.select_all()
            selected = self.chk_frame.get_checked_items()
            self._update_selected_task_summary()

        pricing_mode = self.v_labor_pricing_mode.get() if hasattr(self, "v_labor_pricing_mode") else "avg"

        if pricing_mode == "staff" and not self.project_staff_assignments:
            self.project_staff_last_block_reason = "未导入项目人员名单，已跳过全部专业人工成本"
            for w in self.scroll_frame.winfo_children():
                w.destroy()
            self.dept_rows = {}
            self.refresh_staff_status()
            self.calc_total(show_warning_if_zero=False)
            if not silent_if_missing_staff:
                messagebox.showwarning("提示", "未导入项目人员名单，已自动跳过人工成本，当前仅统计非人工成本。")
            return
        
        agg_days = {}
        task_days = {}
        _, data_params, matrix = self.get_runtime_profile()
        
        try: cap = float(self.e_cap.get())
        except: cap = 100
        
        # --- 国际系数解析逻辑 ---
        intl_str = self.c_intl.get()
        try:
            # 提取括号前的数字，如 "1.2 (国标)" -> 1.2
            intl = float(intl_str.split()[0])
        except:
            intl = 1.0

        selected_canonical = {canonical_task_key(k) for k in selected}
        has_outgoing_task = any(k in OUTGOING_ELIGIBLE_TASKS for k in selected_canonical)
        has_prelim_line_survey_task = any(k in PRELIM_WITH_LINE_SURVEY for k in selected_canonical)

        extras = 0
        if self.v_ext_micro.get(): extras += 6
        if self.v_ext_turbine.get(): extras += 10
        if self.v_ext_proposal.get(): extras += 3
        if self.v_ext_app.get(): extras += 6
        if self.v_ext_bid.get(): extras += 3

        has_outgoing = (self.v_has_outgoing.get() if hasattr(self, "v_has_outgoing") else False) and has_outgoing_task
        has_line_survey = self.v_has_line_survey.get() if hasattr(self, "v_has_line_survey") else False
        is_storage_bundle = (
            (self.v_storage_bundle.get() if hasattr(self, "v_storage_bundle") else False)
            and self.v_type.get() == "储能"
        )
        line_survey_terrain = self.v_line_survey_terrain.get().strip() if hasattr(self, "v_line_survey_terrain") else "非山地"
        terrain_is_mountain = line_survey_terrain.startswith("山")
        # 统一用“地形”口径：选择山地即按复杂地形处理
        is_complex_terrain = terrain_is_mountain
        line_survey_extra = 0.0
        if has_line_survey and has_prelim_line_survey_task:
            line_survey_extra = 10.0 if terrain_is_mountain else 7.0

        booster_voltage = self.v_booster_kv.get().strip() if hasattr(self, "v_booster_kv") else "常规"
        booster_elec_extra_fea_prelim = 0.0
        booster_elec_extra_construction = 0.0
        if booster_voltage.startswith("220"):
            booster_elec_extra_fea_prelim = 2.0
            booster_elec_extra_construction = 5.0
        elif booster_voltage.startswith("330"):
            booster_elec_extra_fea_prelim = 5.0
            booster_elec_extra_construction = 15.0

        has_grid_review = self.v_grid_review.get() if hasattr(self, "v_grid_review") else False

        def add_day(dept_name, day_value, task_name=None):
            agg_days[dept_name] = agg_days.get(dept_name, 0.0) + float(day_value)
            if task_name:
                task_map = task_days.setdefault(task_name, {})
                task_map[dept_name] = task_map.get(dept_name, 0.0) + float(day_value)

        def apply_grid_review_bonus(task_name=None):
            add_day('项目设总', 3.0, task_name)
            add_day('电气一次', 3.0, task_name)
            add_day('电气二次', 3.0, task_name)
        
        for key in selected:
            key = canonical_task_key(key)
            item = data_params.get(key)
            if not item: continue
            
            base = item['base_days']
            extra_days = extras if key in FEA_WITH_FIELD_BOOSTER else 0.0

            factor_cap = get_capacity_factor(cap, item['parsed_rule'])
            
            total_days = (base * factor_cap) + extra_days
            
            depts = matrix.get(key, {})
            if not depts:
                 for k,v in matrix.items(): 
                     if k in key: depts = v; break
            
            if not depts: depts = {"综合": 100}

            for d, r in depts.items():
                add_day(d, total_days * (r / 100.0), key)
            
            # --- 任务附加项加成 ---
            if is_complex_terrain:
                if key in FEA_WITH_FIELD_BOOSTER:
                    add_day('资源', 3.0, key)
                    add_day('线路电气', 3.0, key)
                    add_day('道路', 3.0, key)
                elif key in FEA_STANDALONE_STATION:
                    add_day('总图', 3.0, key)
                if key == "初步设计报告（储能电站）":
                    add_day('总图', 3.0, key)

            # 初设微选附加项（可与对应可研微选重复计入）
            if self.v_prelim_micro_bonus.get():
                prelim_micro_rule = PRELIM_MICRO_BONUS_RULES.get(key)
                if prelim_micro_rule:
                    for dept_name, add_days in (prelim_micro_rule.get("days") or {}).items():
                        add_day(dept_name, float(add_days), key)

            if key in FEA_ALL_BOOSTER_RELATED:
                if booster_elec_extra_fea_prelim > 0:
                    add_day('电气一次', booster_elec_extra_fea_prelim, key)
                    add_day('电气二次', booster_elec_extra_fea_prelim, key)
                if has_grid_review:
                    apply_grid_review_bonus(key)

            # 初设（储能电站）附加项
            if key == "初步设计报告（储能电站）":
                if booster_elec_extra_fea_prelim > 0:
                    add_day('电气一次', booster_elec_extra_fea_prelim, key)
                    add_day('电气二次', booster_elec_extra_fea_prelim, key)

            # 配合电网评审：仅可研与初设增加；初设与对应可研不重复计入
            if has_grid_review and key in PRELIM_ALL_BOOSTER_RELATED:
                related_fea_key = PRELIM_TO_FEA_TASK.get(key)
                if related_fea_key not in selected_canonical:
                    apply_grid_review_bonus(key)

            # 施工图（升压/储能）电压等级附加项
            if key in CONSTRUCTION_STATION_TASKS and booster_elec_extra_construction > 0:
                add_day('电气一次', booster_elec_extra_construction, key)
                add_day('电气二次', booster_elec_extra_construction, key)

            # 外送附加项
            if has_outgoing:
                if key in OUTGOING_FEA_TASKS:
                    add_day('线路电气', 5.0, key)
                    add_day('线路结构', 2.0, key)
                if key in OUTGOING_PRELIM_TASKS:
                    add_day('线路电气', 6.0, key)
                    add_day('线路结构', 3.0, key)
                if key in OUTGOING_CONSTRUCTION_TASKS:
                    add_day('线路电气', 25.0, key)
                    add_day('线路结构', 15.0, key)

            # 线路测绘附加项（初设）
            if line_survey_extra > 0 and key in PRELIM_WITH_LINE_SURVEY:
                add_day('线路电气', line_survey_extra, key)

            # 配储附加项（储能可研/初设）
            if is_storage_bundle and key in STORAGE_BUNDLE_TASKS:
                add_day('综合能源', 8.0, key)
        
        # 国际系数按“总工日 × 系数”统一作用到各专业工日
        if abs(intl - 1.0) > 1e-12:
            for dept_name in list(agg_days.keys()):
                agg_days[dept_name] = agg_days.get(dept_name, 0.0) * intl
            for task_name in list(task_days.keys()):
                for dept_name in list(task_days[task_name].keys()):
                    task_days[task_name][dept_name] = task_days[task_name][dept_name] * intl

        required_depts = [d for d, v in agg_days.items() if abs(float(v)) > 1e-12]
        skipped_depts = set()
        missing_roles = self._validate_required_staff_wages(required_depts) if pricing_mode == "staff" else []
        if pricing_mode == "staff" and missing_roles:
            skipped_depts = {d for d, _ in missing_roles}
            for dept_name in skipped_depts:
                agg_days.pop(dept_name, None)
            for task_name in list(task_days.keys()):
                for dept_name in list(task_days[task_name].keys()):
                    if dept_name in skipped_depts:
                        task_days[task_name].pop(dept_name, None)
            preview = "、".join(sorted(skipped_depts, key=get_dept_sort_index)[:8])
            if len(skipped_depts) > 8:
                preview += f" 等共{len(skipped_depts)}个专业"
            self.project_staff_last_block_reason = f"已跳过缺少有效人员日薪的专业：{preview}"
            self.refresh_staff_status()
            if not silent_if_missing_staff:
                messagebox.showwarning("人员名单不完整", f"以下专业缺少有效人员日薪，已自动跳过人工成本：\n{preview}")

        # 导出用：按勾选任务拆解到专业工时（需与最终纳入成本的专业一致）
        self.task_breakdown_rows = []
        for task_name in selected:
            ckey = canonical_task_key(task_name)
            dept_days_map = task_days.get(ckey, {})
            for dept_name, day_val in sorted(dept_days_map.items(), key=lambda x: get_dept_sort_index(x[0])):
                if abs(float(day_val)) < 1e-12:
                    continue
                self.task_breakdown_rows.append({
                    "task": ckey,
                    "dept": dept_name,
                    "work_days": round(float(day_val), 2),
                    "work_hours": round(float(day_val) * 8.0, 1),
                })

        self.scroll_frame.columnconfigure(0, weight=1); self.scroll_frame.columnconfigure(1, weight=1)
        
        current_ratios = {k: v.get() for k,v in self.vars_ratio.items()} 
        mapped_ratios = {"主设人": current_ratios["主设"], "设计人": current_ratios["设计"], "校核人": current_ratios["校核"], "审核人": current_ratios["审核"]}

        current_depts = set(self.dept_rows.keys())
        target_depts = set(agg_days.keys())
        for stale_dept in current_depts - target_depts:
            stale_row = self.dept_rows.pop(stale_dept, None)
            if stale_row is not None:
                try:
                    stale_row.frame.destroy()
                except Exception:
                    pass

        for i, (d, t) in enumerate(sorted(agg_days.items(), key=lambda x: get_dept_sort_index(x[0]))):
            row = self.dept_rows.get(d)
            if row is None:
                row = DisciplineRow(
                    self.scroll_frame,
                    d,
                    t,
                    mapped_ratios,
                    self.calc_total,
                    wage_overrides=self.project_staff_role_wages.get(d, {}) if pricing_mode == "staff" else {},
                )
                self.dept_rows[d] = row
            else:
                row.apply_profile(
                    t,
                    mapped_ratios,
                    wage_overrides=self.project_staff_role_wages.get(d, {}) if pricing_mode == "staff" else {},
                )
            row.frame.grid(row=i//2, column=i%2, sticky="nsew", padx=5, pady=5)
        
        self.scroll_frame.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        if not skipped_depts:
            self.project_staff_last_block_reason = ""
            self.refresh_staff_status()
        self.calc_total(show_warning_if_zero=False)

    def calc_total(self, show_warning_if_zero=False):
        labor_cost = 0
        total_hours = 0
        
        if self.dept_rows:
            labor_cost = sum(g.get_total_cost() for g in self.dept_rows.values())
            # 重新计算总工时
            total_hours = sum(g.get_total_hours() for g in self.dept_rows.values())
        
        # 重新计算总工日
        total_days = total_hours / 8.0
        
        geo_cost = 0
        if self.geo_panel:
            geo_cost = self.geo_panel.get_cost()
            
        survey_cost = 0
        if self.survey_panel:
            survey_cost = self.survey_panel.get_cost()

        other_cost = 0
        if self.other_panel:
            other_cost = self.other_panel.get_cost()

        base_total = labor_cost + geo_cost + survey_cost + other_cost
        emergency_reserve = base_total * EMERGENCY_RESERVE_RATE
        total = base_total + emergency_reserve

        finance_metrics = {}
        if self.finance_panel:
            finance_metrics = self.finance_panel.update_metrics(labor_cost, geo_cost, survey_cost, other_cost)
        self.last_finance_metrics = finance_metrics or {}
        
        # 修改：恢复工日显示，位于上方；下方显示总成本预估
        txt = (
            f"项目总工日: {total_days:.2f} 天  |  项目总工时: {total_hours:.1f} h\n"
            f"人工: ¥{labor_cost:,.0f} + 地勘: ¥{geo_cost:,.0f} + 测绘: ¥{survey_cost:,.0f} + 其他: ¥{other_cost:,.0f} "
            f"+ 应急储备: ¥{emergency_reserve:,.0f} = ★ 总成本预估: ¥{total:,.0f}"
        )

        if finance_metrics:
            profit_rate = finance_metrics.get("profit_rate")
            profit_text = "--" if profit_rate is None else f"{profit_rate * 100:.2f}%"
            contract_no_tax = finance_metrics.get("contract_no_tax", 0.0)
            finance_total_cost = finance_metrics.get("total_cost", 0.0)
            txt += (
                f"\n利润分析：不含税合同额 ¥{contract_no_tax:,.0f}  |  "
                f"总成本 ¥{finance_total_cost:,.0f}  |  利润率 {profit_text}"
            )
        
        if hasattr(self, 'lbl_total'):
            self.lbl_total.configure(text=txt)

    def _bool_to_text(self, val):
        return "是" if bool(val) else "否"

    def _safe_float(self, val, default=0.0):
        try:
            return float(val)
        except:
            return float(default)

    def _collect_export_config_rows(self):
        rows = []

        def add(section, item, value):
            rows.append({
                "分类": section,
                "项目": item,
                "值": str(value),
            })

        # 1) 基础参数
        add("基础参数", "项目类型", self.v_type.get())
        add("基础参数", "容量(MW)", self.e_cap.get())
        add("基础参数", "国际系数选项", self.c_intl.get())
        add("基础参数", "全局比例-主设(%)", self.vars_ratio["主设"].get())
        add("基础参数", "全局比例-设计(%)", self.vars_ratio["设计"].get())
        add("基础参数", "全局比例-校核(%)", self.vars_ratio["校核"].get())
        add("基础参数", "全局比例-审核(%)", self.vars_ratio["审核"].get())

        # 2) 修订项 + 任务选择
        add("其他修订项", "微观选址(+6)", self._bool_to_text(self.v_ext_micro.get()))
        add("其他修订项", "机型比选(+10)", self._bool_to_text(self.v_ext_turbine.get()))
        add("其他修订项", "建议书(+3)", self._bool_to_text(self.v_ext_proposal.get()))
        add("其他修订项", "项目申请(+6)", self._bool_to_text(self.v_ext_app.get()))
        add("其他修订项", "竞配报告(+3)", self._bool_to_text(self.v_ext_bid.get()))
        add("其他修订项", "配合电网评审", self._bool_to_text(self.v_grid_review.get()))
        add("其他修订项", "含外送", self._bool_to_text(self.v_has_outgoing.get()))
        add("其他修订项", "含线路测绘（仅初设）", self._bool_to_text(self.v_has_line_survey.get()))
        add("其他修订项", "线路测绘地形", self.v_line_survey_terrain.get())
        add("其他修订项", "升压站电压等级", self.v_booster_kv.get())
        add("其他修订项", "若为配储(综合能源+8，仅储能项目适用)", self._bool_to_text(self.v_storage_bundle.get()))
        add("其他修订项", "初设微选附加", self._bool_to_text(self.v_prelim_micro_bonus.get()))

        selected_tasks = [canonical_task_key(x) for x in self.chk_frame.get_checked_items()]
        add("任务选择", "已勾选任务数量", len(selected_tasks))
        if selected_tasks:
            for idx, task_name in enumerate(selected_tasks, start=1):
                add("任务选择", f"任务{idx}", task_name)
        else:
            add("任务选择", "任务", "无")

        # 3) 地勘选定项
        if self.geo_panel:
            g = self.geo_panel
            scale_label = g.e_scale.label_widget.cget("text")
            field_f = self._safe_float(g.e_field_footage.get())
            booster_f = self._safe_float(g.e_booster_footage.get())
            price = self._safe_float(g.e_price.get())
            lab = self._safe_float(g.e_lab_fee.get())
            extra = self._safe_float(g.e_extra_fee.get())
            raw_cost = (field_f + booster_f) * price + lab + extra

            add("地勘采购", "计入总造价", self._bool_to_text(g.v_enable.get()))
            add("地勘采购", "阶段", g.c_stage.get())
            add("地勘采购", "地质", g.c_soil.get())
            add("地勘采购", scale_label, g.e_scale.get())
            add("地勘采购", "含升压站", self._bool_to_text(g.v_booster.get()))
            add("地勘采购", "含储能", self._bool_to_text(g.v_storage.get()))
            add("地勘采购", "含边坡", self._bool_to_text(g.v_slope.get()))
            add("地勘采购", "场区孔数", g.e_holes.get())
            add("地勘采购", "场区进尺(m)", g.e_field_footage.get())
            add("地勘采购", "升压站进尺(m)", g.e_booster_footage.get())
            add("地勘采购", "单价(元/m)", g.e_price.get())
            add("地勘采购", "实验费(元)", g.e_lab_fee.get())
            add("地勘采购", "杂费(元)", g.e_extra_fee.get())
            add("地勘采购", "测算费用(未计入前,元)", round(raw_cost, 2))
            add("地勘采购", "计入金额(元)", round(g.get_cost(), 2))

        # 测绘选定项
        if self.survey_panel:
            s = self.survey_panel
            area = self._safe_float(s.e_area.get())
            if area <= 0:
                unit_price = 0
            elif area <= 30:
                unit_price = 1600
            elif area <= 100:
                unit_price = 1300
            else:
                unit_price = 1100
            raw_cost = area * unit_price

            add("测绘采购", "计入总造价", self._bool_to_text(s.v_enable.get()))
            add("测绘采购", "测绘面积(km²)", s.e_area.get())
            add("测绘采购", "单价(元/km²)", unit_price)
            add("测绘采购", "测算费用(未计入前,元)", round(raw_cost, 2))
            add("测绘采购", "计入金额(元)", round(s.get_cost(), 2))

        # 其他费用选定项
        if self.other_panel:
            o = self.other_panel
            consult_detail = o._get_consult_detail_wan() if o.v_consult.get() else {}
            design_detail = o._get_design_detail_wan() if o.v_design.get() else {}
            consult_wan = sum(consult_detail.values())
            design_wan = sum(design_detail.values())
            raw_cost = (consult_wan + design_wan) * 10000.0
            tier_map = {
                "le_100": "≤100MW",
                "101_200": "101-200MW",
                "201_500": "201-500MW",
                "gt_500": ">500MW",
            }

            add("其他费用", "计入总造价", self._bool_to_text(o.v_enable.get()))
            add("其他费用", "咨询（可研）阶段", self._bool_to_text(o.v_consult.get()))
            add("其他费用", "设计阶段", self._bool_to_text(o.v_design.get()))
            add("其他费用", "设计阶段含驻场", self._bool_to_text(o.v_design_stationed.get()))
            add("其他费用", "图纸国际邮寄", self._bool_to_text(o.v_design_international_mail.get()))
            add("其他费用", "咨询分档", tier_map.get(o._get_consult_tier(), ""))
            add("其他费用", "咨询费用(万元)", round(consult_wan, 4))
            add("其他费用", "设计费用(万元)", round(design_wan, 4))
            add("其他费用", "测算费用(未计入前,元)", round(raw_cost, 2))
            add("其他费用", "计入金额(元)", round(o.get_cost(), 2))

        # 4) 利润分析
        if self.finance_panel:
            m = self.finance_panel.get_metrics()
            add("利润分析", "合同额(含税,元)", round(m.get("contract_taxed", 0.0), 2))
            add("利润分析", "不含税合同额(元)", round(m.get("contract_no_tax", 0.0), 2))
            add("利润分析", "基础成本(人工+外购+差旅,元)", round(m.get("base_cost", 0.0), 2))
            add("利润分析", "应急储备(元)", round(m.get("emergency_reserve", 0.0), 2))
            add("利润分析", "总成本(含应急储备,元)", round(m.get("total_cost", 0.0), 2))
            profit_rate = m.get("profit_rate")
            add("利润分析", "利润率(%)", "" if profit_rate is None else round(profit_rate * 100, 4))

        return rows

    def export_excel(self):
        if HAS_OPENPYXL or HAS_PANDAS:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"成本报价表_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx")
            if not file_path: return
        else:
            file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")], initialfile=f"成本报价表_{datetime.datetime.now().strftime('%Y%m%d')}.csv")
            if not file_path: return

        # 1. 准备数据
        detail_data = []
        detail_export_data = list(getattr(self, "task_breakdown_rows", []) or [])
        config_export_data = self._collect_export_config_rows()
        labor_total = 0
        for dept, row_obj in self.dept_rows.items():
            dept_sum = 0
            for col in row_obj.role_cols:
                cost = col.get_cost()
                dept_sum += cost
                labor_total += cost
                detail_data.append({
                    "专业": dept,
                    "角色": col.role_name,
                    "日薪标准": col.wage,
                    "工日": round(col.hours/8.0, 2),
                    "工时": round(col.hours, 1),
                    "人工成本": cost
                })
                # 任务拆解明细在 generate_report 中计算并缓存到 self.task_breakdown_rows
        
        geo_cost = self.geo_panel.get_cost()
        survey_cost = self.survey_panel.get_cost()
        other_cost = self.other_panel.get_cost() if self.other_panel else 0
        base_total = labor_total + geo_cost + survey_cost + other_cost
        emergency_reserve = base_total * EMERGENCY_RESERVE_RATE
        grand_total = base_total + emergency_reserve

        # 总览表数据 - 修正文案
        summary_data = [
            {"项目": "人工总成本", "金额": labor_total, "备注": "各专业工日汇总"},
            {"项目": "地勘采购", "金额": geo_cost, "备注": f"{self.geo_panel.e_holes.get()}孔 / {self.geo_panel.e_field_footage.get()}m"},
            {"项目": "测绘采购", "金额": survey_cost, "备注": f"{self.survey_panel.e_area.get()} km²"},
            {"项目": "其他费用", "金额": other_cost, "备注": "咨询/设计阶段差旅与打印等"},
            {"项目": "应急储备", "金额": emergency_reserve, "备注": "基础总成本×10%"},
            {"项目": "总成本预估", "金额": grand_total, "备注": "基础总成本×1.1"}
        ]

        if self.finance_panel:
            fm = self.finance_panel.get_metrics()
            profit_rate = fm.get("profit_rate")
            profit_text = "--" if profit_rate is None else f"{profit_rate * 100:.2f}%"
            summary_data.extend(
                [
                    {"项目": "利润测算总成本", "金额": fm.get("total_cost", 0.0), "备注": "含应急储备"},
                    {"项目": "利润率", "金额": 0.0, "备注": profit_text},
                ]
            )

        if HAS_OPENPYXL:
            try:
                export_styled_xlsx(
                    file_path,
                    summary_data,
                    detail_export_data,
                    title="瑞科设计成本核算系统 报价报表",
                    config_data=config_export_data,
                )
                messagebox.showinfo("成功", f"精美报表已导出至：\n{file_path}")
                try:
                    if platform.system() == "Windows":
                        os.startfile(file_path)
                    else:
                        import subprocess
                        subprocess.call(["open", file_path])
                except: pass
            except Exception as e:
                messagebox.showerror("导出失败", str(e))
        elif HAS_PANDAS:
            df_detail = pd.DataFrame(detail_export_data)
            if df_detail.empty:
                df_detail = pd.DataFrame(columns=["任务", "专业", "工日", "工时"])
            else:
                df_detail = df_detail.rename(
                    columns={
                        "task": "任务",
                        "dept": "专业",
                        "work_days": "工日",
                        "work_hours": "工时",
                    }
                )
            df_summary = pd.DataFrame(summary_data)
            df_config = pd.DataFrame(config_export_data)
            if df_config.empty:
                df_config = pd.DataFrame(columns=["分类", "项目", "值"])

            # 2. 写入 Excel (兼容模式)
            try:
                # 优先尝试 xlsxwriter (美化版)
                try:
                    import xlsxwriter
                    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                        # --- Sheet 1: 总价汇总表 ---
                        df_summary.to_excel(writer, sheet_name='总价汇总表', index=False)
                        workbook = writer.book
                        ws_sum = writer.sheets['总价汇总表']
                        
                        # 样式定义
                        fmt_header = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
                        fmt_cell = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
                        fmt_money = workbook.add_format({'border': 1, 'num_format': '¥#,##0', 'align': 'right', 'valign': 'vcenter'})
                        fmt_total = workbook.add_format({'bold': True, 'bg_color': '#FFFF00', 'border': 1, 'num_format': '¥#,##0', 'align': 'right'})

                        # 设置列宽
                        ws_sum.set_column('A:A', 20)
                        ws_sum.set_column('B:B', 20)
                        ws_sum.set_column('C:C', 40)

                        # 应用样式
                        for col_num, value in enumerate(df_summary.columns.values):
                            ws_sum.write(0, col_num, value, fmt_header)
                        
                        for row_num, row_data in enumerate(summary_data):
                            ws_sum.write(row_num+1, 0, row_data['项目'], fmt_cell)
                            style = fmt_total if row_data['项目'] == "总成本预估" else fmt_money
                            ws_sum.write(row_num+1, 1, row_data['金额'], style)
                            ws_sum.write(row_num+1, 2, row_data['备注'], fmt_cell)

                        # --- Sheet 2: 人工明细表 ---
                        df_detail.to_excel(writer, sheet_name='人工明细表', index=False)
                        ws_det = writer.sheets['人工明细表']
                        
                        ws_det.set_column('A:A', 36) # 任务
                        ws_det.set_column('B:B', 15) # 专业
                        ws_det.set_column('C:D', 10) # 工日/工时

                        # 应用表头样式
                        for col_num, value in enumerate(df_detail.columns.values):
                            ws_det.write(0, col_num, value, fmt_header)
                        
                        # 应用数据样式
                        for row in range(1, len(df_detail) + 1):
                            ws_det.write_row(row, 0, df_detail.iloc[row-1], fmt_cell)
                            # 工日/工时列统一为数值格式
                            ws_det.write(row, 2, df_detail.iloc[row-1]['工日'], fmt_cell)
                            ws_det.write(row, 3, df_detail.iloc[row-1]['工时'], fmt_cell)

                        # --- Sheet 3: 参数与选项 ---
                        df_config.to_excel(writer, sheet_name='参数与选项', index=False)
                        ws_cfg = writer.sheets['参数与选项']
                        ws_cfg.set_column('A:A', 18)
                        ws_cfg.set_column('B:B', 28)
                        ws_cfg.set_column('C:C', 90)
                        for col_num, value in enumerate(df_config.columns.values):
                            ws_cfg.write(0, col_num, value, fmt_header)
                        for row in range(1, len(df_config) + 1):
                            ws_cfg.write_row(row, 0, df_config.iloc[row-1], fmt_cell)

                    messagebox.showinfo("成功", f"精美报表已导出至：\n{file_path}")
                
                except ImportError:
                    # 降级处理 (普通版)
                    with pd.ExcelWriter(file_path) as writer:
                        df_summary.to_excel(writer, sheet_name='总价汇总表', index=False)
                        df_detail.to_excel(writer, sheet_name='人工明细表', index=False)
                        df_config.to_excel(writer, sheet_name='参数与选项', index=False)
                    messagebox.showwarning("完成", f"导出成功（未安装xlsxwriter，已生成普通表格）。\n文件位置：{file_path}")
                    
                try:
                    if platform.system() == "Windows":
                        os.startfile(file_path)
                    else:
                        import subprocess
                        subprocess.call(["open", file_path])
                except: pass

            except Exception as e:
                messagebox.showerror("导出失败", str(e))
        else:
            try:
                base = os.path.splitext(file_path)[0]
                file_summary = base + "_总览.csv"
                file_detail = base + "_明细.csv"
                file_config = base + "_参数与选项.csv"
                write_csv(file_summary, ["项目", "金额", "备注"], summary_data)
                detail_csv_rows = [
                    {
                        "任务": x["task"],
                        "专业": x["dept"],
                        "工日": x["work_days"],
                        "工时": x["work_hours"],
                    }
                    for x in detail_export_data
                ]
                write_csv(file_detail, ["任务", "专业", "工日", "工时"], detail_csv_rows)
                write_csv(file_config, ["分类", "项目", "值"], config_export_data)
                messagebox.showinfo("成功", f"CSV 已导出：\n{file_summary}\n{file_detail}\n{file_config}")
            except Exception as e:
                messagebox.showerror("导出失败", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
